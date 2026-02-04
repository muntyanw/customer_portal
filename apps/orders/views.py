# apps/orders/views.py
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django import forms
from django.db import transaction
from django.db.models import Case, DecimalField, ExpressionWrapper, F, Sum, When, Max, Q
from django.db.models.functions import Coalesce
from django.contrib import messages
from .models import (
    Order,
    OrderItem,
    Transaction,
    OrderStatusLog,
    NotificationEmail,
    PaymentMessage,
    OrderComponentItem,
    CurrencyRate,
    CurrencyRateHistory,
    OrderDeletionHistory,
    CurrencyAutoUpdateSettings,
)
from apps.customers.models import CustomerProfile
from apps.accounts.roles import is_manager
import json
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, ROUND_UP
import re
from django.contrib.admin.views.decorators import staff_member_required
from django.conf import settings
from django.core import signing
from django.core.mail import EmailMessage, send_mail
from django.core.files.base import ContentFile
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from django.utils import timezone
import datetime
from .utils_components import parse_components_from_post
from django.urls import reverse
from .services_currency import get_current_eur_rate, update_eur_rate_from_nbu
from django.views.decorators.http import require_POST
from django.http import JsonResponse, Http404, HttpResponse
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.html import format_html, format_html_join, conditional_escape

        
def _get(lst, idx, default=""):
    """EN: Safe list index with default. UA: Безпечне отримання елемента зі списку з дефолтом."""
    return lst[idx] if idx < len(lst) else default


def _to_decimal(value, default="0"):
    """EN: Convert string to Decimal with fallback. UA: Конвертація рядка в Decimal з запасним значенням."""
    value = (value or "").strip()
    if not value:
        value = default
    try:
        return Decimal(value.replace(",", "."))
    except (InvalidOperation, AttributeError):
        return Decimal(default)


def _to_int(value, default=0):
    """EN: Safe int conversion with fallback. UA: Безпечне перетворення в int із запасним значенням."""
    try:
        return int((value or "").strip() or default)
    except (ValueError, TypeError):
        return int(default)


def _normalize_discount_percent(pct):
    try:
        pct = Decimal(pct or 0)
    except Exception:
        pct = Decimal("0")
    if pct < Decimal("-100"):
        pct = Decimal("-100")
    if pct > Decimal("100"):
        pct = Decimal("100")
    return pct.quantize(Decimal("0.01"))


def _customer_discount_multiplier(user=None, pct=None):
    """Return discount multiplier (1 - percent/100). If pct is provided, it takes precedence."""
    if pct is None:
        if not user:
            pct = Decimal("0")
        else:
            profile = getattr(user, "customerprofile", None)
            pct = getattr(profile, "discount_percent", Decimal("0")) if profile else Decimal("0")
    pct = _normalize_discount_percent(pct)
    return (Decimal("100") - pct) / Decimal("100"), pct


def _control_side_label(val):
    """EN: Human label for control side. UA: Людська назва сторони керування."""
    if val == "left":
        return "Ліве"
    if val == "right":
        return "Праве"
    return val or ""


_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})$")
_PROPOSAL_SALT = "orders-proposal-link-v1"
_BALANCE_SALT = "orders-balance-link-v1"
OPTION_LABELS = {
    "magnets_price_eur": "Фіксація магнітами",
    "cord_pvc_tension_price_eur": "Фіксація Леска ПВХ з дотяжкою",
    "cord_copper_barrel_price_eur": "Фіксація Леска з мідною діжкою",
    "magnet_fix_price_eur": "Фіксація магнітами (доп.)",
    "top_pvc_clip_pair_price_eur": "Кліпса кріплення для верхньої планки ПВХ, пара",
    "top_pvc_bar_tape_price_eur_mp": "Верхня планка ПВХ зі скотчем (монтаж без свердління), за м.п.",
    "bottom_wide_bar_price_eur_mp": "Нижня широка планка, за м.п.",
    "top_bar_scotch_price_eur_mp": "Верхній скотч, за м.п.",
    "metal_cord_fix_price_eur": "Металева фіксація шнура",
    "middle_bracket_price_eur": "Проміжковий кронштейн, шт",
    "remote_5ch_price_eur": "5-ти канальний пульт ДУ, шт",
    "remote_15ch_price_eur": "15-ти канальний пульт ДУ, шт",
    "motor_with_remote_price_eur": "Електродвигун з одноканальним пультом ДУ (входить до вартості), шт",
    "motor_no_remote_price_eur": "Електродвигун без пульта, шт",
    "metal_kronsht_price_eur": "Металевий кронштейн, шт",
    "adapter_mosel_internal_price_eur": "Адаптер внутрішній MOSel",
    "adapter_mosel_external_price_eur": "Адаптер зовнішній MOSel",
}


def _proposal_token(order: Order) -> str:
    """Generate a signed token for public commercial proposal links."""
    return signing.dumps({"order": order.pk}, salt=_PROPOSAL_SALT)


def _order_from_token(token: str) -> Order:
    """Return order for a signed token or 404."""
    try:
        data = signing.loads(token, salt=_PROPOSAL_SALT)
    except signing.BadSignature:
        raise Http404("Неправильне посилання.")
    order_id = (data or {}).get("order")
    if not order_id:
        raise Http404("Замовлення не знайдено.")
    return get_object_or_404(Order, pk=order_id, deleted=False)


def _balance_token(customer_id: int) -> str:
    """Generate signed token for public balance page."""
    return signing.dumps({"customer": customer_id}, salt=_BALANCE_SALT)


def _customer_from_balance_token(token: str):
    try:
        data = signing.loads(token, salt=_BALANCE_SALT)
    except signing.BadSignature:
        raise Http404("Невірне посилання.")
    customer_id = (data or {}).get("customer")
    if not customer_id:
        raise Http404("Клієнта не знайдено.")
    return get_object_or_404(get_user_model(), pk=customer_id, is_customer=True)


def _collect_item_options(item: OrderItem, rate: Decimal, markup_multiplier: Decimal = Decimal("1")):
    """
    Build list of option rows for public proposal with EUR/UAH.
    """
    options = []
    rate_with_markup = Decimal(rate or 0) * Decimal(markup_multiplier or 1)
    for field in [
        f.name
        for f in OrderItem._meta.fields
        if f.name.endswith("_price_eur") or f.name.endswith("_price_eur_mp")
    ]:
        price = getattr(item, field, None)
        if price in (None, ""):
            continue
        qty_field = field.replace("_price_eur_mp", "_qty").replace("_price_eur", "_qty")
        qty_val = getattr(item, qty_field, 0) or 0
        try:
            price_val = Decimal(price)
            qty_val_dec = Decimal(qty_val)
        except Exception:
            continue
        if price_val <= 0 or qty_val_dec <= 0:
            continue
        options.append(
            {
                "label": OPTION_LABELS.get(field, field),
                "qty": qty_val_dec,
                "price_eur": price_val,
                "price_uah": _round_uah_total(price_val * rate_with_markup),
            }
        )
    return options


class CurrencyAutoUpdateForm(forms.Form):
    auto_update = forms.BooleanField(
        required=False,
        label="Оновлювати курс автоматично",
        widget=forms.CheckboxInput(attrs={"class": "form-check-input"}),
    )
    update_times = forms.CharField(
        required=False,
        label="Часи оновлення",
        widget=forms.TextInput(
            attrs={
                "class": "form-control",
                "placeholder": "08:30, 12:00, 16:30",
            }
        ),
        help_text="Формат HH:MM, кілька значень через кому або пробіл.",
    )

    def clean_update_times(self):
        raw = (self.cleaned_data.get("update_times") or "").strip()
        if not raw:
            return []

        raw = raw.replace(";", ",").replace("\u00a0", " ")
        parts = re.split(r"[,\\s]+", raw)
        times = []
        seen = set()
        for part in parts:
            part = part.strip()
            if not part:
                continue
            if "." in part and ":" not in part:
                part = part.replace(".", ":")
            match = _TIME_RE.match(part)
            if not match:
                raise forms.ValidationError("Час має бути у форматі HH:MM (наприклад, 09:30).")
            hour = int(match.group(1))
            minute = int(match.group(2))
            if hour > 23 or minute > 59:
                raise forms.ValidationError("Час має бути у форматі HH:MM (наприклад, 09:30).")
            normalized = f"{hour:02d}:{minute:02d}"
            if normalized in seen:
                continue
            seen.add(normalized)
            times.append(normalized)

        return times


def _orders_scope(user):
    if is_manager(user):
        return Order.objects.filter(deleted=False)
    return Order.objects.filter(customer=user, deleted=False)


def _parse_date_range(params):
    """
    Read date_from/date_to from GET params, fallback to last 30 days ending today.
    Returns (date_from, date_to, date_from_str, date_to_str).
    """
    date_from_str = (params.get("date_from") or "").strip()
    date_to_str = (params.get("date_to") or "").strip()
    date_from = None
    date_to = None

    if date_from_str:
        try:
            date_from = datetime.datetime.strptime(date_from_str, "%Y-%m-%d").date()
        except ValueError:
            date_from = None
    if date_to_str:
        try:
            date_to = datetime.datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            date_to = None

    today = timezone.localdate()
    if not date_to:
        date_to = today
        date_to_str = date_to.isoformat()
    if not date_from:
        date_from = date_to - datetime.timedelta(days=30)
        date_from_str = date_from.isoformat()

    return date_from, date_to, date_from_str, date_to_str


def _transactions_scope(user):
    if is_manager(user):
        return Transaction.objects.select_related("customer", "customer__customerprofile", "created_by", "order").filter(deleted=False)
    return Transaction.objects.select_related("customer", "customer__customerprofile", "created_by", "order").filter(customer=user, deleted=False)

def _order_rate(order, current_rate: Decimal) -> Decimal:
    """
    Pick rate: frozen for робота+, або поточний для прорахунку.
    """
    if order.status != Order.STATUS_QUOTE and order.eur_rate:
        return Decimal(order.eur_rate)
    return Decimal(current_rate or 0)

def _round_uah_total(value: Decimal) -> Decimal:
    # Округлюємо завжди вгору (ceil) для узгодження з фронтом
    return Decimal(value or 0).quantize(Decimal("1"), rounding=ROUND_UP)


def _order_base_total(order) -> Decimal:
    """
    Return stored total (без націнки). Якщо відсутнє/нульове, рахуємо по items.
    """
    stored = Decimal(order.total_eur or 0)
    if stored > 0:
        return stored
    # subtotal_eur already includes quantity from the builder UI
    agg = order.items.aggregate(
        total=Sum("subtotal_eur")
    ).get("total") or Decimal("0")
    return agg

STATUS_LABELS = dict(Order.STATUS_CHOICES)
STATUS_BADGES = {
    Order.STATUS_QUOTE: "secondary",
    Order.STATUS_IN_WORK: "warning",
    Order.STATUS_READY: "info",
    Order.STATUS_SHIPPED: "success",
}


def _set_order_totals_uah(orders, current_rate: Decimal):
    """Attach total_uah_display/display_rate to orders for templates."""
    for o in orders:
        rate = _order_rate(o, current_rate)
        total_eur = _order_base_total(o)
        o.display_rate = rate
        extra_uah = Decimal(getattr(o, "extra_service_amount_uah", 0) or 0)
        # Display без націнки, але з додатковою послугою
        o.total_uah_display = _round_uah_total(total_eur * rate + extra_uah)
    return orders


def _order_total_uah(order, current_rate: Decimal) -> Decimal:
    extra_uah = Decimal(getattr(order, "extra_service_amount_uah", 0) or 0)
    total = _order_base_total(order) * _order_rate(order, current_rate) * (Decimal("1") + Decimal(order.markup_percent or 0) / Decimal("100")) + extra_uah
    return _round_uah_total(total)


def _orders_total_uah(orders_qs, current_rate: Decimal) -> Decimal:
    return sum(
        _round_uah_total(_order_base_total(o) * _order_rate(o, current_rate) * (Decimal("1") + Decimal(o.markup_percent or 0) / Decimal("100")) + Decimal(getattr(o, "extra_service_amount_uah", 0) or 0))
        for o in orders_qs
    )


def _orders_total_uah_base(orders_qs, current_rate: Decimal) -> Decimal:
    """Base total for balances: без націнок і без додаткових послуг."""
    return sum(
        _round_uah_total(_order_base_total(o) * _order_rate(o, current_rate))
        for o in orders_qs
    )


def _transactions_total_uah(tx_qs) -> Decimal:
    if tx_qs is None:
        return Decimal("0")
    tx_total = tx_qs.aggregate(
        total=Sum(
            Case(
                When(
                    type=Transaction.DEBIT,
                    then=ExpressionWrapper(
                        F("amount") * F("eur_rate"),
                        output_field=DecimalField(max_digits=18, decimal_places=4),
                    ),
                ),
                default=ExpressionWrapper(
                    -F("amount") * F("eur_rate"),
                    output_field=DecimalField(max_digits=18, decimal_places=4),
                ),
                output_field=DecimalField(max_digits=18, decimal_places=4),
            )
        )
    )["total"] or Decimal("0")
    return Decimal(tx_total).quantize(Decimal("0.01"))

def _tx_amount_uah(tx: Transaction) -> Decimal:
    """Return transaction amount in UAH with sign (debit +, credit -)."""
    amount = Decimal(tx.amount or 0)
    rate = Decimal(tx.eur_rate or 0)
    total = amount * rate
    if tx.type == Transaction.CREDIT:
        total = -total
    return Decimal(total).quantize(Decimal("0.01"))


def _get_payment_message_text():
    return (
        PaymentMessage.objects.filter(is_active=True)
        .order_by("-created_at")
        .values_list("text", flat=True)
        .first()
    )


def _payment_shortage_context(order):
    """Return dict with shortage and list of orders covering it (for prompt)."""
    if not order:
        return None
    current_rate = get_current_eur_rate()
    balance = compute_balance(order.customer)
    shortage = Decimal(balance or 0)
    if shortage >= 0:
        return None
    shortage = shortage * -1

    cover_orders = []
    cover_total = Decimal("0")
    existing_orders = (
        _orders_scope(order.customer)
        .filter(status__in=[Order.STATUS_IN_WORK, Order.STATUS_READY, Order.STATUS_SHIPPED])
        .exclude(pk=order.pk)
        .order_by("-created_at")
    )
    if order.pk:
        cover_orders.append(order.pk)
        cover_total += _order_total_uah(order, current_rate)
    for o in existing_orders:
        cover_total += _order_total_uah(o, current_rate)
        cover_orders.append(o.pk)
        if cover_total >= shortage:
            break

    return {
        "shortage": shortage.quantize(Decimal("0.01")),
        "orders": ([order.pk] if order.pk else []) + cover_orders,
    }


def _build_order_workbook(order, save_to_file: bool = True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Замовлення"
    profile = getattr(order.customer, "customerprofile", None)
    rate = Decimal(order.eur_rate or get_current_eur_rate() or 0)
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_horiz = Border(top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")
    center_wrap_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def price_uah(eur, qty=1):
        if not rate:
            return float(Decimal(eur or 0) * Decimal(qty or 0))
        return float((Decimal(eur or 0) * Decimal(qty or 0) * rate).quantize(Decimal("1"), rounding=ROUND_UP))

    row = 1
    client_info = [
        ("Клієнт", getattr(profile, "full_name", "") or str(order.customer)),
        ("Тел.", getattr(profile, "phone", "") or ""),
        ("ПІБ.", getattr(profile, "full_name", "") or ""),
        ("Адреса доставки", ""),
        ("Примітки до замовлення", order.note or ""),
    ]
    for label, val in client_info:
        ws.append(["", label, val])
    # widen column B and center
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["D"].width = 20
    for r in range(1, ws.max_row + 1):
        ws.cell(row=r, column=2).alignment = center_align

    # merge C-L in rows 1-5 and add borders B-L, center text
    for r in range(1, 6):
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.border = border_all
        ws.cell(row=r, column=3).alignment = center_wrap_align

    # пустая строка после блоку приміток
    ws.append([])

    total_order_uah = Decimal("0")
    item_total_rows = []
    gap_rows = []
    ws.append([""] * 12)
    total_row = ws.max_row
    ws.merge_cells(start_row=total_row, start_column=7, end_row=total_row, end_column=11)
    ws.cell(row=total_row, column=7).value = "Загальна вартість замовлення, грн"
    ws.cell(row=total_row, column=7).alignment = center_wrap_align
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    for c in range(7, 12 + 1):
        ws.cell(row=total_row, column=c).border = border_all
    total_cell = ws.cell(row=total_row, column=12)
    row = ws.max_row + 1

    created_str = order.created_at.astimezone(timezone.get_current_timezone()).strftime("%d.%m.%y") if order.created_at else ""

    def control_label(val):
        if val == "left":
            return "Ліве"
        if val == "right":
            return "Праве"
        return val or ""

    system_colors = [
        "біла",
        "коричнева",
        "графіт",
        "золотий дуб",
        "сіра",
    ]

    def system_color(val):
        if not val:
            return val
        lower = str(val).lower()
        for col in system_colors:
            if col in lower:
                return col
        return val

    def add_option_rows(item):
        rows = []
        total_opts = Decimal("0")
        for field in [
            f.name
            for f in OrderItem._meta.fields
            if f.name.endswith("_price_eur") or f.name.endswith("_price_eur_mp")
        ]:
            price = getattr(item, field, None)
            qty_field = field.replace("_price_eur_mp", "_qty").replace("_price_eur", "_qty")
            qty = getattr(item, qty_field, None)
            if price in (None, "") or qty in (None, ""):
                continue
            try:
                qty_val = Decimal(qty)
                price_val = Decimal(price)
            except Exception:
                continue
            if qty_val <= 0 or price_val <= 0:
                continue
            label = OPTION_LABELS.get(field, field)
            # builder stores total price per option (already * qty), keep as is
            sum_uah = price_uah(price_val)
            rows.append(["", label, "", "", "", "", "", "", "", "", float(qty_val), sum_uah])
            total_opts += Decimal(sum_uah)
        return rows, total_opts

    headers = ["", "Система", "Колір с-ми", "Тканина", "Колір тканини", "Ширина, мм", "Знач. Шир.", "Висота, мм", "Знач. Вис.", "Сторона управ.", "К-сть", "Вартість, грн"]

    for idx, it in enumerate(order.items.all(), start=1):
        start_item_row = ws.max_row + 1
        ws.append([f"Поз. {idx}", f"Замовлення № {order.pk} від {created_str}"])
        title_row = ws.max_row
        ws.cell(row=title_row, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=title_row, start_column=2, end_row=title_row, end_column=12)
        ws.cell(row=title_row, column=2).alignment = center_wrap_align
        ws.cell(row=title_row, column=2).font = Font(bold=True, size=14)

        ws.append(headers)
        header_row = ws.max_row
        for c in range(2, 13):
            cell = ws.cell(row=header_row, column=c)
            cell.font = Font(bold=True)
            cell.alignment = center_wrap_align

        # subtotal_eur already includes quantity (див. _order_base_total)
        total_item_uah = Decimal(str(price_uah(it.subtotal_eur)))
        opt_rows, total_opts = add_option_rows(it)

        ws.append([
            "",
            it.system_sheet,
            system_color(it.table_section),
            it.fabric_name,
            it.fabric_color_code,
            it.width_fabric_mm,
            "По тканині" if it.gabarit_width_flag else "",
            it.height_gabarit_mm,
            "По тканині" if it.fabric_height_flag else (it.roll_height_info or "Габарит"),
            control_label(it.control_side),
            it.quantity,
            float(max(total_item_uah - Decimal(total_opts), Decimal("0"))),
        ])
        # center align parameters row
        param_row = ws.max_row
        for c in range(2, 13):
            ws.cell(row=param_row, column=c).alignment = center_align

        if opt_rows:
            start_opt_row = ws.max_row + 1
            for r in opt_rows:
                label = r[1]
                qty_val = r[10] if len(r) > 10 else ""
                price_val = r[11] if len(r) > 11 else ""
                ws.append([""] * 12)
                cur_row = ws.max_row
                # опция: название в C-K, цена в L
                ws.merge_cells(start_row=cur_row, start_column=3, end_row=cur_row, end_column=11)
                label_cell = ws.cell(row=cur_row, column=3)
                label_cell.value = f"{label} ({qty_val})" if qty_val not in ("", None) else label
                label_cell.alignment = Alignment(vertical="center")
                for c in range(2, 13):
                    ws.cell(row=cur_row, column=c).border = border_all
                ws.cell(row=cur_row, column=12).value = price_val
            end_opt_row = ws.max_row
            # Додатково по вертикалі в C
            ws.merge_cells(start_row=start_opt_row, start_column=2, end_row=end_opt_row, end_column=2)
            extra_cell = ws.cell(row=start_opt_row, column=2)
            extra_cell.value = "Додатково:"
            extra_cell.alignment = center_align
            for r in range(start_opt_row, end_opt_row + 1):
                ws.cell(row=r, column=2).border = border_all

            ws.append(["", "", "", "", "", "", "", "", "", "Всього:", it.quantity, float(total_item_uah)])
            item_total_rows.append(ws.max_row)
        else:
            ws.append(["", "", "", "", "", "", "", "", "", "Всього:", it.quantity, float(total_item_uah)])
            item_total_rows.append(ws.max_row)

        if it.note:
            ws.append(["", "Примітки до виробу", it.note] + [""] * 9)
            note_row = ws.max_row
            ws.merge_cells(start_row=note_row, start_column=3, end_row=note_row, end_column=12)
            ws.cell(row=note_row, column=3).alignment = center_align
        if it.gabarit_width_flag:
            gabarit_note = it.roll_height_info or "Габаритна ширина"
            ws.append(["", "Примітка габаритної ширини", gabarit_note] + [""] * 9)
            gabarit_row = ws.max_row
            ws.merge_cells(start_row=gabarit_row, start_column=3, end_row=gabarit_row, end_column=12)
            ws.cell(row=gabarit_row, column=3).alignment = center_align
        # пустая строка после блока позиции без боковых границ
        ws.append([""] * 12)
        gap_rows.append(ws.max_row)

        end_item_row = ws.max_row
        ws.merge_cells(start_row=start_item_row, start_column=1, end_row=end_item_row, end_column=1)
        ws.cell(row=start_item_row, column=1).alignment = center_align
        ws.cell(row=start_item_row, column=1).border = Border(left=thin, right=thin)

        total_order_uah += total_item_uah

    if order.component_items.exists():
        ws.append(["", "Комплектуючі"])
        ws.append(["", "Найменування", "Колір", "Од. вим", "К-сть", "Ціна, грн"])
        for comp in order.component_items.all():
            price_uah_val = price_uah(comp.price_eur, comp.quantity)
            ws.append(["", comp.name, comp.color, comp.unit, float(comp.quantity or 0), price_uah_val])
            total_order_uah += Decimal(price_uah_val)
        ws.append([])

    total_cell.value = float(_round_uah_total(total_order_uah)) if total_order_uah else 0

    # center all column B cells
    for r in range(1, ws.max_row + 1):
        ws.cell(row=r, column=2).alignment = center_wrap_align

    # add borders to all populated grid cells (A-L)
    for r in range(1, ws.max_row + 1):
        for c in range(1, 13):
            if r in gap_rows:
                if c == 1:
                    ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                else:
                    ws.cell(row=r, column=c).border = border_horiz
            elif c == 1:
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin)
            elif c == 12:
                ws.cell(row=r, column=c).border = Border(right=thin, left=thin, top=thin, bottom=thin)
            else:
                ws.cell(row=r, column=c).border = border_all

    # override borders: row 6 horizontal only, before total labels remove verticals
    # blank row after header (row 6)
    for c in range(1, 7):
        ws.cell(row=6, column=c).border = Border()
    for c in range(7, 13):
        ws.cell(row=6, column=c).border = border_horiz
    ws.cell(row=6, column=1).border = Border(left=thin)
    # overall total row: remove verticals before text (cols 1-6)
    for c in range(1, 7):
        ws.cell(row=total_row, column=c).border = border_horiz
    # item totals: remove verticals before "Всього:"
    for r in item_total_rows:
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = border_horiz
        ws.cell(row=r, column=1).border = Border(left=thin, right=thin)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"order_{order.pk}.xlsx"
    if save_to_file:
        order.workbook_file.save(filename, ContentFile(buffer.getvalue()), save=False)
    return filename, buffer.getvalue()


@login_required
def order_workbook_download(request, pk):
    """Managers: always regenerate and download the Excel workbook for an order."""
    if not is_manager(request.user):
        raise Http404
    order = get_object_or_404(Order, pk=pk)
    filename, data = _build_order_workbook(order, save_to_file=False)
    response = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_proposal_workbook(order):
    # NOTE: This function used to delegate to _render_order_workbook(), but that helper
    # is not guaranteed to exist. Keep proposal generation self-contained here.
    wb = Workbook()
    ws = wb.active
    ws.title = "Пропозиція"
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # light yellow
    position_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # light green

    rate = _order_rate(order, get_current_eur_rate())
    base_total_eur = _order_base_total(order)
    markup = Decimal(order.markup_percent or 0)
    markup_multiplier = Decimal("1") + markup / Decimal("100")
    extra_uah = Decimal(getattr(order, "extra_service_amount_uah", 0) or 0)
    total_with_markup_uah = _round_uah_total(
        base_total_eur * markup_multiplier * Decimal(rate or 0) + extra_uah
    )
    created_str = (
        order.created_at.astimezone(timezone.get_current_timezone()).strftime("%d.%m.%Y %H:%M")
        if order.created_at
        else ""
    )
    profile = getattr(order.customer, "customerprofile", None)
    customer_name = getattr(profile, "company_name", "") or getattr(profile, "full_name", "") or str(order.customer)
    customer_phone = getattr(profile, "phone", "") or ""
    customer_full_name = getattr(profile, "full_name", "") or str(order.customer)
    customer_address = getattr(profile, "trade_address", "") or ""

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 60

    ws.append([f"Комерційна пропозиція №{order.pk}", ""])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).fill = header_fill
    if created_str:
        ws.append(["Створено", created_str])
    ws.append(["Клієнт", customer_name])
    ws.append(["Телефон", customer_phone])
    ws.append(["ПІБ", customer_full_name])
    ws.append(["Адреса", customer_address])
    if extra_uah:
        ws.append(["Додаткова послуга", f"{order.extra_service_label or 'Послуга'} — {float(extra_uah)} грн"])
    ws.append(["Вартість замовлення, грн", float(total_with_markup_uah)])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.append([])

    # форматування перших рядків
    for r in range(2, 8):
        if ws.cell(row=r, column=1).value is not None:
            ws.cell(row=r, column=1).font = Font(bold=True)
        if ws.cell(row=r, column=2).value is not None:
            ws.cell(row=r, column=2).alignment = center_align

    def add_row(label, value):
        ws.append([label, value])
        lbl = (label or "").lower()
        if lbl.startswith("вартість") or lbl.startswith("сума") or lbl.startswith("ціна"):
            ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    def control_label(val):
        if val == "left":
            return "Ліве"
        if val == "right":
            return "Праве"
        return val or ""

    for idx, it in enumerate(order.items.all(), start=1):
        ws.append([f"Позиція {idx}", ""])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.cell(row=ws.max_row, column=1).alignment = center_align
        ws.cell(row=ws.max_row, column=1).fill = position_fill
        fabric = it.fabric_name or ""
        if it.fabric_color_code:
            fabric = f"{fabric} ({it.fabric_color_code})" if fabric else it.fabric_color_code

        add_row("Система", it.system_sheet)
        add_row("Колір системи", it.table_section)
        add_row("Тканина", fabric)
        add_row("Ширина, мм", it.width_fabric_mm)
        add_row("Висота, мм", it.height_gabarit_mm)
        add_row("Кількість", int(it.quantity or 0))
        unit_price = Decimal(it.subtotal_eur or 0)
        try:
            qty_dec = Decimal(it.quantity or 1)
            if qty_dec > 0:
                unit_price = unit_price / qty_dec
        except Exception:
            pass
        unit_uah = _round_uah_total(unit_price * Decimal(rate or 0) * markup_multiplier)
        add_row("Ціна за 1 шт, грн", float(unit_uah))
        add_row("Сторона керування", control_label(it.control_side))
        subtotal_uah = _round_uah_total(Decimal(it.subtotal_eur or 0) * Decimal(rate or 0) * markup_multiplier)
        add_row("Вартість, грн", float(subtotal_uah))
        if it.note:
            add_row("Примітка", it.note)

        options = _collect_item_options(it, rate, markup_multiplier)
        if options:
            ws.append(["Додаткові опції", ""])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
            ws.cell(row=ws.max_row, column=1).alignment = center_align
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            for opt in options:
                qty_raw = opt.get("qty") or ""
                qty = int(qty_raw) if isinstance(qty_raw, (int, float, Decimal)) else qty_raw
                price_uah = opt.get("price_uah") or Decimal("0")
                detail = f"{qty} · {price_uah} грн" if qty not in ("", None) else f"{price_uah} грн"
                add_row(opt.get("label", ""), detail)
        ws.append([])

    if order.component_items.exists():
        ws.append(["Комплектуючі"])
        ws.append(["Характеристика", "Значення"])
        for comp in order.component_items.all():
            total_uah = _round_uah_total(
                Decimal(comp.price_eur or 0) * Decimal(comp.quantity or 0) * Decimal(rate or 0) * markup_multiplier
            )
            add_row("Найменування", comp.name)
            if comp.color:
                add_row("Колір", comp.color)
            if comp.unit:
                add_row("Од. вим.", comp.unit)
            add_row("Кількість", float(comp.quantity or 0))
            add_row("Сума, грн", float(total_uah))
            ws.append([])

    # Borders for filled cells + стили колонок (жирный/центр)
    max_r = ws.max_row
    for r in range(1, max_r + 1):
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = border_all
            if c == 1 and ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).font = Font(bold=True)
            if c == 2 and ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).alignment = center_align

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"proposal_{order.pk}.xlsx"
    return filename, buffer.getvalue()


def _send_order_in_work_email(order, request, workbook_payload=None):
    """Send notification to configured recipients when an order enters 'in work'."""
    recipients = list(
        NotificationEmail.objects.filter(is_active=True).values_list("email", flat=True)
    )
    if not recipients:
        return

    profile = getattr(order.customer, "customerprofile", None)
    full_name = getattr(profile, "full_name", "") or ""
    phone = getattr(profile, "phone", "") or ""

    try:
        order_url = request.build_absolute_uri(reverse("orders:builder_edit", args=[order.pk]))
    except Exception:
        order_url = ""

    subject = f"Замовлення #{order.pk} відправлено в роботу"
    body_lines = [
        f"Замовлення #{order.pk} відправлено в роботу.",
        f"Клієнт: {order.customer}",
    ]
    if full_name:
        body_lines.append(f"ПІБ: {full_name}")
    if phone:
        body_lines.append(f"Телефон: {phone}")
    total_uah = getattr(order, "total_uah_display", "")
    if not total_uah:
        try:
            total_uah = _round_uah_total(Decimal(order.total_eur or 0) * Decimal(order.eur_rate or 0))
        except Exception:
            total_uah = ""
    if total_uah != "":
        body_lines.append(f"Сума (UAH): {total_uah}")
    if order_url:
        body_lines.append(f"Посилання: {order_url}")

    email = EmailMessage(
        subject=subject,
        body="\n".join(body_lines),
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=recipients,
    )

    filename = None
    data = None
    if workbook_payload:
        filename, data = workbook_payload
    elif order.workbook_file:
        try:
            with order.workbook_file.open("rb") as f:
                data = f.read()
                filename = order.workbook_file.name.split("/")[-1]
        except Exception:
            data = None
    if filename and data:
        email.attach(filename, data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    email.send(fail_silently=True)


def _prepare_to_work(order, request):
    """Set rates for quote->work and block if credit not allowed."""
    if order.status != Order.STATUS_QUOTE:
        return True

    current_rate = get_current_eur_rate()
    order.eur_rate = current_rate
    order.eur_rate_at_creation = current_rate
    order.save(update_fields=["eur_rate", "eur_rate_at_creation"])

    balance_before = compute_balance(order.customer)
    order_total_uah = _round_uah_total(_order_base_total(order) * Decimal(current_rate or 0))
    projected_balance = balance_before - order_total_uah
    shortage = projected_balance * -1 if projected_balance < 0 else Decimal("0")
    try:
        profile = order.customer.customerprofile
    except CustomerProfile.DoesNotExist:
        profile = None
    credit_allowed = bool(profile and profile.credit_allowed)
    if shortage > 0 and not credit_allowed and not is_manager(request.user):
        messages.warning(
            request,
            f"Оплатіть суму {shortage} для відправки в роботу.",
        )
        return False
    return True


def _apply_status_change(order, new_status, request):
    """Shared status transition logic for rollers & components."""
    if not new_status or new_status == order.status:
        return True

    workbook_payload = None
    if new_status == Order.STATUS_IN_WORK:
        if not _prepare_to_work(order, request):
            return False
        workbook_payload = _build_order_workbook(order)
        order.status = new_status
        order.save(update_fields=["status", "workbook_file"])
        OrderStatusLog.objects.create(
            order=order,
            status=new_status,
            user=request.user,
        )
        _send_order_in_work_email(order, request, workbook_payload)
        return True

    order.status = new_status
    order.save(update_fields=["status"])
    OrderStatusLog.objects.create(
        order=order,
        status=new_status,
        user=request.user,
    )
    return True


def compute_balance(user, force_personal: bool = False):
    current_rate = get_current_eur_rate()
    if force_personal:
        orders_qs = Order.objects.filter(customer=user, deleted=False)
        tx_qs = Transaction.objects.filter(customer=user, deleted=False)
    else:
        orders_qs = _orders_scope(user)
        tx_qs = _transactions_scope(user)

    orders_qs = orders_qs.filter(
        status__in=[
            Order.STATUS_IN_WORK,
            Order.STATUS_READY,
            Order.STATUS_SHIPPED,
        ]
    )
    orders_sum_uah = _orders_total_uah_base(orders_qs, current_rate)
    tx_total_uah = _transactions_total_uah(tx_qs)
    return tx_total_uah - orders_sum_uah


class TransactionForm(forms.ModelForm):
    orders = forms.ModelMultipleChoiceField(
        queryset=Order.objects.none(),
        required=True,
        label="Замовлення",
        widget=forms.SelectMultiple(attrs={"class": "form-select", "size": "6"}),
    )

    class Meta:
        model = Transaction
        fields = ["customer", "type", "amount", "description", "payment_type", "account_number"]
        widgets = {
            "description": forms.Textarea(attrs={"rows": 3}),
            "payment_type": forms.Select(attrs={"class": "form-select"}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.current_rate = get_current_eur_rate()
        rate_decimal = Decimal(self.current_rate or 0)
        self.current_rate = rate_decimal.quantize(Decimal("0.0001")) if rate_decimal else rate_decimal
        User = get_user_model()
        self.fields["customer"].queryset = (
            User.objects.select_related("customerprofile")
            .order_by("customerprofile__full_name", "email")
        )

        def _label(u):
            profile = getattr(u, "customerprofile", None)
            full_name = getattr(profile, "full_name", "") or u.get_full_name() or u.username
            phone = getattr(profile, "phone", "") or ""
            contact_email = getattr(profile, "contact_email", "") or u.email
            parts = [full_name]
            if phone:
                parts.append(phone)
            if contact_email:
                parts.append(contact_email)
            return " · ".join(parts)

        self.fields["customer"].label_from_instance = _label
        self.fields["customer"].widget.attrs.update({"class": "form-select", "data-customer-picker": "true"})
        self.fields["type"].widget.attrs.update({"class": "form-select"})
        self.fields["orders"].queryset = Order.objects.filter(deleted=False)
        self.fields["orders"].widget.attrs.update(
            {
                "class": "form-select",
                "data-order-picker": "true",
                "multiple": "multiple",
                "size": "8",
            }
        )
        self.fields["amount"].label = "Сума, грн"
        self.fields["payment_type"].label = "Вид оплати"
        self.fields["account_number"].label = "Номер рахунку (для 'На рахунок')"
        self.fields["account_number"].widget.attrs.update({"placeholder": "UA1234567890...", "class": "form-control"})
        self.fields["amount"].widget.attrs.update(
            {"class": "form-control", "step": "0.01", "min": "0", "inputmode": "decimal"}
        )
        self.fields["description"].widget.attrs.update({"class": "form-control"})

    def clean(self):
        cleaned = super().clean()
        ptype = cleaned.get("payment_type")
        acc = (cleaned.get("account_number") or "").strip()
        if not cleaned.get("orders"):
            raise forms.ValidationError("Оберіть хоча б одне замовлення.")
        if ptype == Transaction.PAY_ACCOUNT and not acc:
            raise forms.ValidationError("Для оплати на рахунок потрібно вказати номер рахунку.")
        if ptype != Transaction.PAY_ACCOUNT:
            cleaned["account_number"] = ""
        return cleaned

    def clean_amount(self):
        amount_uah = self.cleaned_data.get("amount") or Decimal("0")
        rate = Decimal(self.current_rate or 0)
        if rate <= 0:
            raise forms.ValidationError("Немає актуального курсу EUR для конвертації.")
        return (amount_uah / rate).quantize(Decimal("0.00001"))

    def save(self, commit=True):
        orders = list(self.cleaned_data.get("orders") or [])
        order_for_link = orders[0] if orders else None
        desc = (self.cleaned_data.get("description") or "").strip()
        if orders:
            ids = ", ".join(str(o.pk) for o in orders)
            if desc:
                desc = f"{desc} (Замовлення: {ids})"
            else:
                desc = f"Замовлення: {ids}"

        obj = Transaction(
            customer=self.cleaned_data.get("customer"),
            type=self.cleaned_data.get("type"),
            amount=self.cleaned_data.get("amount"),
            eur_rate=Decimal(self.current_rate or 0),
            description=desc,
            payment_type=self.cleaned_data.get("payment_type"),
            account_number=self.cleaned_data.get("account_number") or "",
            order=order_for_link,
        )
        if commit:
            obj.save()
        # збережемо вибрані замовлення, щоб вюха могла оновити їх нотатки
        self._selected_orders = orders
        return obj

@login_required
def order_list(request):
    """
    EN: List of roller orders (with positions).
    UA: Список замовлень з тканинними ролетами.
    """
    User = get_user_model()
    status_filter = request.GET.get("status") or ""
    customer_filter = request.GET.get("customer") or ""
    q = (request.GET.get("q") or "").strip()
    date_from, date_to, date_from_str, date_to_str = _parse_date_range(request.GET)
    balance_user = request.user
    orders_qs = (
        _orders_scope(request.user)
        .select_related("customer", "customer__customerprofile")
        .annotate(last_status_at=Coalesce(Max("status_logs__created_at"), "created_at"))
        .prefetch_related("status_logs")
        .exclude(component_items__isnull=False)  # exclude component orders from rollers list
        .order_by("-id")
        .distinct()
    )

    if status_filter:
        orders_qs = orders_qs.filter(status=status_filter)
    if date_from:
        orders_qs = orders_qs.filter(created_at__date__gte=date_from)
    if date_to:
        orders_qs = orders_qs.filter(created_at__date__lte=date_to)

    if customer_filter and is_manager(request.user):
        balance_user = get_object_or_404(User, pk=customer_filter)
        orders_qs = orders_qs.filter(customer=balance_user)
    if q:
        orders_qs = orders_qs.filter(pk__icontains=q)

    customers_filter_list = (
        User.objects.filter(orders__isnull=False).select_related("customerprofile").distinct()
        if is_manager(request.user) else []
    )
    current_rate = get_current_eur_rate()
    orders_qs = _set_order_totals_uah(orders_qs, current_rate)
    quote_orders_count = orders_qs.filter(status=Order.STATUS_QUOTE).count()
    show_bulk_select = quote_orders_count > 1
    payment_message_text = _get_payment_message_text() or ""
    payment_shortage_map = {}
    proposal_tokens = {o.id: _proposal_token(o) for o in orders_qs}
    proposal_page_urls = {oid: reverse("orders:proposal_page", args=[tok]) for oid, tok in proposal_tokens.items()}
    proposal_excel_urls = {oid: reverse("orders:proposal_excel", args=[tok]) for oid, tok in proposal_tokens.items()}
    if quote_orders_count:
        for o in orders_qs:
            if o.status != Order.STATUS_QUOTE or o.customer_id != request.user.id:
                continue
            shortage_ctx = _payment_shortage_context(o)
            if shortage_ctx:
                payment_shortage_map[o.id] = {
                    "shortage": str(shortage_ctx.get("shortage") or ""),
                    "orders": shortage_ctx.get("orders") or [],
                }

    context = {
        "orders": orders_qs,
        "statuses": Order.STATUS_CHOICES,
        "status_filter": status_filter,
        "customer_filter": customer_filter,
        "date_from": date_from_str,
        "date_to": date_to_str,
        "customer_options": customers_filter_list,
        "list_mode": "rollers",
        "status_badges": STATUS_BADGES,
        "status_labels": STATUS_LABELS,
        "quote_orders_count": quote_orders_count,
        "show_bulk_select": show_bulk_select,
        "q": q,
        "payment_message_text": payment_message_text,
        "payment_shortage_map": payment_shortage_map,
        "payment_shortage_json": json.dumps(payment_shortage_map),
        "user_balance": compute_balance(balance_user),
        "proposal_page_urls": proposal_page_urls,
        "proposal_excel_urls": proposal_excel_urls,
    }
    return render(request, "orders/order_list.html", context)


@login_required
def order_components_list(request):
    """
    EN: List of orders that have components.
    UA: Список замовлень, у яких є комплектуючі.
    """
    User = get_user_model()
    status_filter = request.GET.get("status") or ""
    customer_filter = request.GET.get("customer") or ""
    q = (request.GET.get("q") or "").strip()
    date_from, date_to, date_from_str, date_to_str = _parse_date_range(request.GET)
    balance_user = request.user

    qs = (
        _orders_scope(request.user)
        .select_related("customer", "customer__customerprofile")
        .annotate(last_status_at=Coalesce(Max("status_logs__created_at"), "created_at"))
        .prefetch_related("status_logs")
        .filter(component_items__isnull=False)
        .distinct()
        .order_by("-id")
    )

    if status_filter:
        qs = qs.filter(status=status_filter)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if customer_filter and is_manager(request.user):
        balance_user = get_object_or_404(User, pk=customer_filter)
        qs = qs.filter(customer=balance_user)
    if q:
        qs = qs.filter(pk__icontains=q)

    customers_filter_list = (
        User.objects.filter(orders__isnull=False).select_related("customerprofile").distinct()
        if is_manager(request.user) else []
    )
    current_rate = get_current_eur_rate()
    qs = _set_order_totals_uah(qs, current_rate)
    quote_orders_count = qs.filter(status=Order.STATUS_QUOTE).count()
    show_bulk_select = quote_orders_count > 1
    payment_message_text = _get_payment_message_text() or ""
    payment_shortage_map = {}
    proposal_tokens = {o.id: _proposal_token(o) for o in qs}
    proposal_page_urls = {oid: reverse("orders:proposal_page", args=[tok]) for oid, tok in proposal_tokens.items()}
    proposal_excel_urls = {oid: reverse("orders:proposal_excel", args=[tok]) for oid, tok in proposal_tokens.items()}
    if quote_orders_count:
        for o in qs:
            if o.status != Order.STATUS_QUOTE or o.customer_id != request.user.id:
                continue
            shortage_ctx = _payment_shortage_context(o)
            if shortage_ctx:
                payment_shortage_map[o.id] = {
                    "shortage": str(shortage_ctx.get("shortage") or ""),
                    "orders": shortage_ctx.get("orders") or [],
                }

    context = {
        "orders": qs,
        "list_mode": "components",
        "statuses": Order.STATUS_CHOICES,
        "status_filter": status_filter,
        "customer_filter": customer_filter,
        "date_from": date_from_str,
        "date_to": date_to_str,
        "customer_options": customers_filter_list,
        "status_badges": STATUS_BADGES,
        "status_labels": STATUS_LABELS,
        "quote_orders_count": quote_orders_count,
        "show_bulk_select": show_bulk_select,
        "q": q,
        "payment_message_text": payment_message_text,
        "payment_shortage_map": payment_shortage_map,
        "payment_shortage_json": json.dumps(payment_shortage_map),
        "user_balance": compute_balance(balance_user),
        "proposal_page_urls": proposal_page_urls,
        "proposal_excel_urls": proposal_excel_urls,
    }
    return render(request, "orders/order_list.html", context)


@login_required
def order_notifications_settings(request):
    """
    EN: Manage recipient emails for order-in-work notifications.
    UA: Керування email-адресами для сповіщень про відправку замовлення в роботу.
    """
    if not is_manager(request.user):
        messages.error(request, "Доступно лише менеджерам.")
        return redirect("orders:list")

    EmailFormSet = forms.modelformset_factory(
        NotificationEmail,
        fields=["email", "is_active"],
        extra=1,
        can_delete=True,
        widgets={
            "email": forms.EmailInput(attrs={"class": "form-control", "placeholder": "email@example.com"}),
            "is_active": forms.CheckboxInput(attrs={"class": "form-check-input"}),
        },
    )
    PaymentFormSet = forms.modelformset_factory(
        PaymentMessage,
        fields=["text", "is_active"],
        extra=1,
        can_delete=True,
        widgets={
            "text": forms.Textarea(attrs={"class": "form-control", "rows": 2, "placeholder": "Текст повідомлення"}),
            "is_active": forms.CheckboxInput(attrs={"class": "form-check-input"}),
        },
    )

    emails_qs = NotificationEmail.objects.order_by("email")
    payments_qs = PaymentMessage.objects.order_by("-created_at")
    currency_settings = CurrencyAutoUpdateSettings.get_solo()

    if request.method == "POST":
        email_formset = EmailFormSet(request.POST, queryset=emails_qs, prefix="emails")
        payment_formset = PaymentFormSet(request.POST, queryset=payments_qs, prefix="payments")
        currency_form = CurrencyAutoUpdateForm(request.POST, prefix="currency")
        if email_formset.is_valid() and payment_formset.is_valid() and currency_form.is_valid():
            email_formset.save()
            payment_formset.save()
            currency_settings.auto_update = currency_form.cleaned_data["auto_update"]
            currency_settings.update_times = currency_form.cleaned_data["update_times"]
            currency_settings.save(update_fields=["auto_update", "update_times"])
            messages.success(request, "Налаштування збережено.")
            return redirect("orders:notifications_settings")
    else:
        email_formset = EmailFormSet(queryset=emails_qs, prefix="emails")
        payment_formset = PaymentFormSet(queryset=payments_qs, prefix="payments")
        currency_form = CurrencyAutoUpdateForm(
            prefix="currency",
            initial={
                "auto_update": currency_settings.auto_update,
                "update_times": ", ".join(currency_settings.update_times or []),
            },
        )

    return render(
        request,
        "orders/notification_settings.html",
        {
            "email_formset": email_formset,
            "payment_formset": payment_formset,
            "currency_form": currency_form,
        },
    )


@login_required
def order_detail(request, pk: int):
    return redirect("orders:builder_edit", pk=pk)


@login_required
def order_update(request, pk: int):
    """
    Customer може редагувати лише свої замовлення (за потреби — тільки певні поля).
    Manager — будь-які.
    """
    if is_manager(request.user):
        order = get_object_or_404(Order, pk=pk, deleted=False)
    else:
        order = get_object_or_404(Order, pk=pk, customer=request.user, deleted=False)

    if request.method == "POST":
        form = OrderForm(request.POST, request.FILES, instance=order)
        if form.is_valid():
            form.save()
            return redirect("orders:list")
    else:
        form = OrderForm(instance=order)
    return render(request, "orders/update.html", {"form": form, "order": order})


PRICE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1vjwqhZ0-9SWcN-u8Oa-T6ciNmHfMeHU-c2RTv6axqHs/edit?gid=0#gid=0"


@login_required
@transaction.atomic
def order_builder(request, pk=None):
    """
    Один универсальный билдер:
    - GET + pk=None     → создание заказа (пустой билдер)
    - GET + pk=<id>     → редактирование / просмотр
    - POST + pk=None    → создать заказ + Items
    - POST + pk=<id>    → обновить заказ + Items
    """

    # Получение или создание ордера
    if pk:
        if is_manager(request.user):
            order = get_object_or_404(Order, pk=pk, deleted=False)
        else:
            order = get_object_or_404(Order, pk=pk, customer=request.user, deleted=False)
    else:
        order = None

    readonly = bool(order and (not is_manager(request.user) and order.status != Order.STATUS_QUOTE))

    # -------------------------------------
    # POST: создание или обновление
    # -------------------------------------
    if request.method == "POST":
        if readonly:
            messages.warning(request, "Це замовлення можна лише переглядати.")
            return redirect("orders:builder_edit", pk=order.pk)
        action = request.POST.get("status_action") or "save"
        chosen_customer = None
        can_pick_customer = is_manager(request.user) or request.user.is_staff or request.user.is_superuser
        if can_pick_customer:
            cust_id = request.POST.get("customer_id")
            if cust_id:
                try:
                    chosen_customer = get_user_model().objects.get(pk=cust_id)
                except get_user_model().DoesNotExist:
                    chosen_customer = None
            if chosen_customer is None:
                messages.error(request, "Оберіть клієнта для замовлення.")
                return redirect("orders:builder")
        target_customer = chosen_customer or (order.customer if order else request.user)
        if order and not chosen_customer:
            discount_pct_val = _normalize_discount_percent(order.discount_percent)
        else:
            _, discount_pct_val = _customer_discount_multiplier(target_customer)
        discount_multiplier, _ = _customer_discount_multiplier(pct=discount_pct_val)
        # Если ордера нет — создаём
        if order is None:
            markup_percent = _to_decimal(request.POST.get("markup_percent"), default="0")
            current_rate = get_current_eur_rate()
            order = Order.objects.create(
                customer=target_customer,
                eur_rate_at_creation=current_rate,
                eur_rate=current_rate,
                markup_percent=markup_percent,
                discount_percent=discount_pct_val,
                status=Order.STATUS_QUOTE,
            )
        elif chosen_customer and can_pick_customer:
            order.customer = chosen_customer
        base_note = (request.POST.get("note") or "").strip()
        if chosen_customer and can_pick_customer:
            base_note = re.sub(r"\s*\(створено менеджером [^)]+\)\s*$", "", base_note).strip()
            suffix = f" (створено менеджером {request.user.email})"
            base_note = (base_note + suffix).strip()
        order.note = base_note
        order.extra_service_label = (request.POST.get("extra_service_label") or "").strip()
        extra_service_amount_uah = _to_decimal(request.POST.get("extra_service_amount_uah"), default="0")

        # Для менеджера можно добавить присвоение customer
        profile = getattr(request.user, "customerprofile", None)
        org = getattr(profile, "organization", None)
        update_fields = ["markup_percent", "eur_rate", "total_eur", "eur_rate_at_creation", "note", "extra_service_label", "extra_service_amount_uah", "discount_percent"]

        # Чтение всех списков (как раньше)
        systems = request.POST.getlist("system_sheet")
        sections = request.POST.getlist("table_section")
        fabrics = request.POST.getlist("fabric_name")
        fabric_colors = request.POST.getlist("fabric_color_code")

        h_list = request.POST.getlist("height_gabarit_mm")
        w_list = request.POST.getlist("width_fabric_mm")

        gw_states = request.POST.getlist("gabarit_width_flag_state")
        if gw_states and len(gw_states) < len(systems):
            gw_states += [""] * (len(systems) - len(gw_states))
        gw_flags = request.POST.getlist("gabarit_width_flag")
        gh_flags = request.POST.getlist("fabric_height_flag")
        GbDiffWidthMm = request.POST.getlist("GbDiffWidthMm")
        gb_width_mm = request.POST.getlist("gb_width_mm")

        base_prices = request.POST.getlist("base_price_eur")
        sur_prices = request.POST.getlist("surcharge_height_eur")
        magnets_price_eur = request.POST.getlist("magnets_price_eur")
        magnets_qty = request.POST.getlist("magnets_qty")
        cord_pvc_tension_price_eur = request.POST.getlist("cord_pvc_tension_price_eur")
        cord_pvc_tension_qty = request.POST.getlist("cord_pvc_tension_qty")
        cord_copper_barrel_price_eur = request.POST.getlist("cord_copper_barrel_price_eur")
        cord_copper_barrel_qty = request.POST.getlist("cord_copper_barrel_qty")
        top_pvc_clip_pair_price_eur = request.POST.getlist("top_pvc_clip_pair_price_eur")
        top_pvc_clip_pair_qty = request.POST.getlist("top_pvc_clip_pair_qty")
        top_pvc_bar_tape_price_eur_mp = request.POST.getlist("top_pvc_bar_tape_price_eur_mp")
        top_pvc_bar_tape_qty = request.POST.getlist("top_pvc_bar_tape_qty")
        bottom_wide_bar_price_eur_mp = request.POST.getlist("bottom_wide_bar_price_eur_mp")
        bottom_wide_bar_qty = request.POST.getlist("bottom_wide_bar_qty")
        top_bar_scotch_price_eur_mp = request.POST.getlist("top_bar_scotch_price_eur_mp")
        top_bar_scotch_qty = request.POST.getlist("top_bar_scotch_qty")
        metal_cord_fix_price_eur = request.POST.getlist("metal_cord_fix_price_eur")
        metal_cord_fix_qty = request.POST.getlist("metal_cord_fix_qty")
        middle_bracket_price_eur = request.POST.getlist("middle_bracket_price_eur")
        middle_bracket_qty = request.POST.getlist("middle_bracket_qty")
        remote_15ch_price_eur = request.POST.getlist("remote_15ch_price_eur")
        remote_15ch_qty = request.POST.getlist("remote_15ch_qty")
        remote_5ch_price_eur = request.POST.getlist("remote_5ch_price_eur")
        remote_5ch_qty = request.POST.getlist("remote_5ch_qty")
        motor_with_remote_price_eur = request.POST.getlist("motor_with_remote_price_eur")
        motor_with_remote_qty = request.POST.getlist("motor_with_remote_qty")
        motor_no_remote_price_eur = request.POST.getlist("motor_no_remote_price_eur")
        motor_no_remote_qty = request.POST.getlist("motor_no_remote_qty")
        metal_kronsht_price_eur = request.POST.getlist("metal_kronsht_price_eur")
        metal_kronsht_qty = request.POST.getlist("metal_kronsht_qty")
        
        subtotals = request.POST.getlist("subtotal_eur")

        roll_infos = request.POST.getlist("roll_height_info")
        qty_list = request.POST.getlist("quantity")
        control_sides = request.POST.getlist("control_side")
        bottom_fixations = request.POST.getlist("bottom_fixation")
        pvc_planks = request.POST.getlist("pvc_plank")
        item_notes = request.POST.getlist("item_note")

        # Если нет ни одной позиции — не трогаем существующие items и возвращаем с ошибкой
        if not any((system or "").strip() for system in systems):
            messages.error(request, "Додайте хоча б одну позицію перед відправкою.")
            return redirect("orders:builder_edit", pk=order.pk)

        # Удаляем старые Items только после валидации наличия новых
        order.items.all().delete()

        # -----------------------------------------
        # Создание всех OrderItem (ГОРАЗДО ЧИЩЕ)
        # -----------------------------------------
        for idx, system_sheet in enumerate(systems):
            OrderItem.objects.create(
                order=order,
                organization=org,
                system_sheet=system_sheet or "",
                table_section=_get(sections, idx, ""),
                fabric_name=_get(fabrics, idx, ""),
                fabric_color_code=_get(fabric_colors, idx, ""),
                height_gabarit_mm=_to_int(_get(h_list, idx, "0"), 0),
                width_fabric_mm=_to_int(_get(w_list, idx, "0"), 0),
                gabarit_width_flag=(
                    _get(gw_states, idx) in ("1", "true", "on")
                    if gw_states
                    else _get(gw_flags, idx) in ("on", "true", "1")
                ),
                fabric_height_flag=_get(gh_flags, idx) in ("on", "true", "1"),
                base_price_eur=_to_decimal(_get(base_prices, idx)),
                gb_width_mm=_to_decimal(_get(gb_width_mm, idx)),
                GbDiffWidthMm=_to_decimal(_get(GbDiffWidthMm, idx)),
                surcharge_height_eur=_to_decimal(_get(sur_prices, idx)),
                magnets_price_eur=_to_decimal(_get(magnets_price_eur, idx)),
                magnets_qty=_to_decimal(_get(magnets_qty, idx)),
                cord_pvc_tension_price_eur=_to_decimal(_get(cord_pvc_tension_price_eur, idx)),
                cord_pvc_tension_qty=_to_decimal(_get(cord_pvc_tension_qty, idx)),
                cord_copper_barrel_price_eur=_to_decimal(_get(cord_copper_barrel_price_eur, idx)),
                cord_copper_barrel_qty=_to_decimal(_get(cord_copper_barrel_qty, idx)),
                top_pvc_clip_pair_price_eur=_to_decimal(_get(top_pvc_clip_pair_price_eur, idx)),
                top_pvc_clip_pair_qty=_to_decimal(_get(top_pvc_clip_pair_qty, idx)),
                top_pvc_bar_tape_price_eur_mp=_to_decimal(_get(top_pvc_bar_tape_price_eur_mp, idx)),
                top_pvc_bar_tape_qty=_to_decimal(_get(top_pvc_bar_tape_qty, idx)),
                bottom_wide_bar_price_eur_mp=_to_decimal(_get(bottom_wide_bar_price_eur_mp, idx)),
                bottom_wide_bar_qty=_to_decimal(_get(bottom_wide_bar_qty, idx)),
                top_bar_scotch_price_eur_mp=_to_decimal(_get(top_bar_scotch_price_eur_mp, idx)),
                top_bar_scotch_qty=_to_decimal(_get(top_bar_scotch_qty, idx)),
                metal_cord_fix_price_eur=_to_decimal(_get(metal_cord_fix_price_eur, idx)),
                metal_cord_fix_qty=_to_decimal(_get(metal_cord_fix_qty, idx)),
                
                middle_bracket_price_eur = _to_decimal(_get(middle_bracket_price_eur, idx)),
                middle_bracket_qty = _to_decimal(_get(middle_bracket_qty, idx)),
                remote_15ch_price_eur=_to_decimal(_get(remote_15ch_price_eur, idx)),
                remote_15ch_qty=_to_decimal(_get(remote_15ch_qty, idx)),
                remote_5ch_price_eur=_to_decimal(_get(remote_5ch_price_eur, idx)),
                remote_5ch_qty=_to_decimal(_get(remote_5ch_qty, idx)),
                motor_with_remote_price_eur=_to_decimal(_get(motor_with_remote_price_eur, idx)),
                motor_with_remote_qty=_to_decimal(_get(motor_with_remote_qty, idx)),
                motor_no_remote_price_eur=_to_decimal(_get(motor_no_remote_price_eur, idx)),
                motor_no_remote_qty=_to_decimal(_get(motor_no_remote_qty, idx)),
                metal_kronsht_price_eur=_to_decimal(_get(metal_kronsht_price_eur, idx)),
                metal_kronsht_qty=_to_decimal(_get(metal_kronsht_qty, idx)),
        
                subtotal_eur=_to_decimal(_get(subtotals, idx)).quantize(Decimal("0.01")),
                roll_height_info=_get(roll_infos, idx, ""),
                quantity=_to_int(_get(qty_list, idx, "1"), 1),
                control_side=_get(control_sides, idx, "").strip(),
                bottom_fixation=_get(bottom_fixations, idx) in ("on", "true", "1"),
                pvc_plank=_get(pvc_planks, idx) in ("on", "true", "1"),
                note=_get(item_notes, idx, "").strip(),
            )

        markup_percent = _to_decimal(request.POST.get("markup_percent"), default=str(order.markup_percent or "0"))
        eur_rate_value = _to_decimal(
            request.POST.get("eur_rate"),
            default=str(order.eur_rate or get_current_eur_rate()),
        )

        items_total = order.items.aggregate(
            total=Sum("subtotal_eur")
        ).get("total") or Decimal("0")

        total_with_markup = items_total * (Decimal("1") + markup_percent / Decimal("100"))

        order.extra_service_amount_uah = extra_service_amount_uah.quantize(Decimal("0.01"))
        order.markup_percent = markup_percent.quantize(Decimal("0.01"))
        order.eur_rate = eur_rate_value
        if not order.eur_rate_at_creation:
            order.eur_rate_at_creation = eur_rate_value
        order.total_eur = items_total.quantize(Decimal("0.01"))
        order.discount_percent = discount_pct_val
        if chosen_customer and can_pick_customer:
            order.customer = chosen_customer
            update_fields.append("customer")
        order.save(update_fields=update_fields)

        new_status = order.status
        if action == "to_work":
            new_status = Order.STATUS_IN_WORK
        elif action == "next" and is_manager(request.user):
            new_status = order.next_status() or order.status
        elif action == "prev" and is_manager(request.user):
            new_status = order.prev_status() or order.status

        if new_status and not _apply_status_change(order, new_status, request):
            return redirect("orders:builder_edit", pk=order.pk)

        messages.success(request, "Замовлення збережено.")
        next_url = (request.POST.get("next_url") or "").strip()
        if next_url and url_has_allowed_host_and_scheme(
            next_url,
            allowed_hosts={request.get_host()},
            require_https=request.is_secure(),
        ):
            return redirect(next_url)
        return redirect("orders:list")

    # -----------------------------------------
    # GET: рендер билдер (создание/редактирование)
    # -----------------------------------------

    customers_filter_list = []
    if is_manager(request.user) or request.user.is_staff or request.user.is_superuser:
        customers_filter_list = list(
            CustomerProfile.objects.select_related("user")
            .order_by("full_name", "user__email")
        )

    # Если новый заказ
    if order is None:
        items_json = "[]"
        next_status_label = STATUS_LABELS.get(Order.STATUS_IN_WORK)
    else:
        # экспорт Items для JS
        items = []
        for it in order.items.all().order_by("id"):
            items.append({
                "system_sheet": it.system_sheet,
                "table_section": it.table_section,
                "fabric_name": it.fabric_name,
                "fabric_color_code": it.fabric_color_code,
                "height_gabarit_mm": it.height_gabarit_mm,
                "width_fabric_mm": it.width_fabric_mm,
                "gabarit_width_flag": it.gabarit_width_flag,
                "fabric_height_flag": it.fabric_height_flag,
                "base_price_eur": float(it.base_price_eur),
                "GbDiffWidthMm": float(it.GbDiffWidthMm),
                "gb_width_mm": float(it.gb_width_mm),
                "surcharge_height_eur": float(it.surcharge_height_eur),
                
                "magnets_price_eur": float(it.magnets_price_eur),
                "magnets_qty": float(it.magnets_qty),
                "cord_pvc_tension_price_eur": float(it.cord_pvc_tension_price_eur),
                "cord_pvc_tension_qty": float(it.cord_pvc_tension_qty),
                "cord_copper_barrel_price_eur": float(it.cord_copper_barrel_price_eur),
                "cord_copper_barrel_qty": float(it.cord_copper_barrel_qty),
                "top_pvc_clip_pair_price_eur": float(it.top_pvc_clip_pair_price_eur),
                "top_pvc_clip_pair_qty": float(it.top_pvc_clip_pair_qty),
                "top_pvc_bar_tape_price_eur_mp": float(it.top_pvc_bar_tape_price_eur_mp),
                "top_pvc_bar_tape_qty": float(it.top_pvc_bar_tape_qty),
                "bottom_wide_bar_price_eur_mp": float(it.bottom_wide_bar_price_eur_mp),
                "bottom_wide_bar_qty": float(it.bottom_wide_bar_qty),
                "top_bar_scotch_price_eur_mp": float(it.top_bar_scotch_price_eur_mp),
                "top_bar_scotch_qty": float(it.top_bar_scotch_qty),
                "metal_cord_fix_price_eur": float(it.metal_cord_fix_price_eur),
                "metal_cord_fix_qty": float(it.metal_cord_fix_qty),
                
                "middle_bracket_price_eur": float(it.middle_bracket_price_eur),
                "middle_bracket_qty": float(it.middle_bracket_qty),
                "remote_15ch_price_eur": float(it.remote_15ch_price_eur),
                "remote_15ch_qty": float(it.remote_15ch_qty),
                "remote_5ch_price_eur": float(it.remote_5ch_price_eur),
                "remote_5ch_qty": float(it.remote_5ch_qty),
                "motor_with_remote_price_eur": float(it.motor_with_remote_price_eur),
                "motor_with_remote_qty": float(it.motor_with_remote_qty),
                "motor_no_remote_price_eur": float(it.motor_no_remote_price_eur),
                "motor_no_remote_qty": float(it.motor_no_remote_qty),
                "metal_kronsht_price_eur": float(it.metal_kronsht_price_eur),
                "metal_kronsht_qty": float(it.metal_kronsht_qty),
                
                "subtotal_eur": float(it.subtotal_eur),
                "quantity": it.quantity,
                "roll_height_info": it.roll_height_info,
                "control_side": it.control_side,
                "bottom_fixation": it.bottom_fixation,
                "pvc_plank": it.pvc_plank,
                "note": it.note,
            })
        items_json = json.dumps(items)
        next_code = order.next_status()
        next_status_label = STATUS_LABELS.get(next_code) if next_code else None

    status_logs = order.status_logs.all() if order else []

    builder_rate = get_current_eur_rate()
    if order and order.status != Order.STATUS_QUOTE and order.eur_rate:
        builder_rate = order.eur_rate

    if order:
        discount_percent = _normalize_discount_percent(getattr(order, "discount_percent", Decimal("0")))
    else:
        discount_user = request.user if not (is_manager(request.user) or request.user.is_staff or request.user.is_superuser) else None
        discount_percent = _normalize_discount_percent(
            getattr(getattr(discount_user, "customerprofile", None), "discount_percent", Decimal("0")) if discount_user else Decimal("0")
        )

    customer_discount_percent_js = format(discount_percent, "f")

    shortage_ctx = _payment_shortage_context(order)
    payment_prompt = request.session.pop("payment_prompt", None)
    proposal_token = _proposal_token(order) if order else None
    proposal_page_url = reverse("orders:proposal_page", args=[proposal_token]) if proposal_token else None
    proposal_excel_url = reverse("orders:proposal_excel", args=[proposal_token]) if proposal_token else None

    return render(request, "orders/builder.html", {
        "order": order,
        "items_json": items_json,
        "PRICE_SHEET_URL": PRICE_SHEET_URL,
        "status_logs": status_logs,
        "builder_rate": builder_rate,
        "next_status_label": next_status_label,
        "readonly": readonly,
        "payment_message_text": (payment_prompt or {}).get("text") or _get_payment_message_text(),
        "payment_shortage": payment_prompt or shortage_ctx,
        "status_badges": STATUS_BADGES,
        "proposal_page_url": proposal_page_url,
        "proposal_excel_url": proposal_excel_url,
        "customers_filter_list": customers_filter_list,
        "customer_discount_percent": discount_percent,
        "customer_discount_percent_js": customer_discount_percent_js,
    })
    
    
@login_required
def order_delete(request, pk: int):
    """
    Удаление заказа.
    Customer — может удалить только свой.
    Manager — может удалить любой.
    """
    redirect_target = "orders:list"
    if is_manager(request.user):
        order = get_object_or_404(Order, pk=pk, deleted=False)
    else:
        order = get_object_or_404(Order, pk=pk, customer=request.user, deleted=False)

    # Ограничение: клієнт може видаляти лише статус "Прорахунок"
    if (not is_manager(request.user)) and order.status != Order.STATUS_QUOTE:
        messages.warning(request, "Ви можете видаляти лише замовлення в статусі 'Прорахунок'.")
        return redirect(redirect_target if redirect_target == "orders:list" else reverse(redirect_target))

    # если есть комплектующие, возвращаем на список комплектующих
    if order.component_items.exists():
        redirect_target = "orders:components_list"

    if request.method == "POST":
        order.deleted = True
        order.deleted_at = timezone.now()
        if request.user.is_authenticated:
            order.deleted_by = request.user
        order.save(update_fields=["deleted", "deleted_at", "deleted_by"])
        messages.success(request, "Замовлення переміщено до кошика.")
        return redirect(redirect_target)

    # GET → страница підтвердження
    return render(request, "orders/delete_confirm.html", {"order": order, "soft": True})


@login_required
def order_trash_list(request):
    qs = Order.objects.filter(deleted=True).select_related("customer", "customer__customerprofile")
    if not is_manager(request.user):
        qs = qs.filter(customer=request.user)
    qs = qs.order_by("-deleted_at", "-id")
    return render(
        request,
        "orders/trash_list.html",
        {"orders": qs, "is_manager": is_manager(request.user)},
    )


@login_required
def order_restore(request, pk: int):
    if is_manager(request.user):
        order = get_object_or_404(Order, pk=pk, deleted=True)
    else:
        order = get_object_or_404(Order, pk=pk, customer=request.user, deleted=True)

    if request.method == "POST":
        order.deleted = False
        order.deleted_at = None
        order.deleted_by = None
        order.save(update_fields=["deleted", "deleted_at", "deleted_by"])
        messages.success(request, "Замовлення відновлено.")
        return redirect("orders:trash")

    return render(request, "orders/delete_confirm.html", {"order": order, "restore": True})


@login_required
def transaction_create(request):
    if not is_manager(request.user):
        messages.error(request, "Створювати транзакції може лише менеджер.")
        return redirect("orders:list")

    current_rate = get_current_eur_rate()
    if Decimal(current_rate or 0) <= 0:
        messages.warning(request, "Немає актуального курсу EUR. Оновіть курс перед створенням транзакції.")

    initial = {}
    if "customer" in request.GET:
        initial["customer"] = request.GET.get("customer")

    if request.method == "POST":
        form = TransactionForm(request.POST)
        if form.is_valid():
            tx = form.save(commit=False)
            tx.created_by = request.user
            tx.save()

            orders_selected = getattr(form, "_selected_orders", []) or []
            # Оновлюємо опис транзакції посиланнями на замовлення
            if orders_selected:
                order_links = tuple(
                    (request.build_absolute_uri(reverse("orders:builder_edit", args=[o.pk])), o.pk)
                    for o in orders_selected
                )
                base_desc = conditional_escape(form.cleaned_data.get("description") or "")
                links_html = format_html_join(", ", '<a href="{0}">#{1}</a>', order_links)
                combined = format_html(
                    "{}{}{}",
                    base_desc,
                    format_html(" — ") if base_desc and links_html else "",
                    format_html("Замовлення: {0}", links_html) if links_html else "",
                )
                tx.description = combined
                tx.save(update_fields=["description"])

                # Додаємо примітку до замовлень з номером транзакції та лінком
                tx_link = request.build_absolute_uri(reverse("orders:balances")) + f"?customer={tx.customer_id}"
                tx_anchor = format_html('<a href="{0}">Транзакція №{1}</a>', tx_link, tx.pk)
                for o in orders_selected:
                    existing = o.note or ""
                    append = format_html("{}", tx_anchor)
                    if existing:
                        new_note = format_html("{}\n{}", existing, append)
                    else:
                        new_note = append
                    o.note = new_note
                    o.save(update_fields=["note"])

            messages.success(request, "Транзакцію створено.")
            return redirect("orders:balances")
    else:
        form = TransactionForm(initial=initial)

    return render(request, "orders/transaction_form.html", {"form": form, "current_rate": current_rate})


@login_required
def transaction_delete(request, pk: int):
    if not is_manager(request.user):
        messages.error(request, "Видаляти транзакції може лише менеджер.")
        return redirect("orders:balances")
    tx = get_object_or_404(Transaction, pk=pk, deleted=False)
    if request.method == "POST":
        tx.deleted = True
        tx.deleted_at = timezone.now()
        tx.deleted_by = request.user
        tx.save(update_fields=["deleted", "deleted_at", "deleted_by"])
        messages.success(request, f"Транзакцію №{tx.pk} переміщено до кошика.")
        return redirect("orders:balances")
    return render(request, "orders/transaction_delete_confirm.html", {"tx": tx})


@login_required
def transaction_trash(request):
    if not is_manager(request.user):
        messages.error(request, "Доступ заборонено.")
        return redirect("orders:balances")
    txs = Transaction.objects.filter(deleted=True).select_related("customer", "customer__customerprofile", "deleted_by").order_by("-deleted_at")
    return render(request, "orders/transaction_trash.html", {"transactions": txs})


@login_required
def transaction_restore(request, pk: int):
    if not is_manager(request.user):
        messages.error(request, "Доступ заборонено.")
        return redirect("orders:balances")
    tx = get_object_or_404(Transaction, pk=pk, deleted=True)
    if request.method == "POST":
        tx.deleted = False
        tx.deleted_at = None
        tx.deleted_by = None
        tx.save(update_fields=["deleted", "deleted_at", "deleted_by"])
        messages.success(request, f"Транзакцію №{tx.pk} відновлено.")
        return redirect("orders:transaction_trash")
    return render(request, "orders/transaction_restore_confirm.html", {"tx": tx})


@login_required
def balances_history(request):
    """
    EN: Orders + transactions timeline with balance and filters.
    UA: Історія замовлень і транзакцій з фільтрами.
    """
    User = get_user_model()
    status_filter = request.GET.get("status") or ""
    customer_filter = request.GET.get("customer") or ""
    type_filter = request.GET.get("type") or ""
    negative_only = request.GET.get("negative") in ("1", "true", "on")
    balance_user = request.user

    orders_qs = (
        _orders_scope(request.user)
        .select_related("customer", "customer__customerprofile")
        .prefetch_related("status_logs", "component_items")
        .exclude(status=Order.STATUS_QUOTE)  # Прорахунок не впливає на баланс
        .order_by("-created_at")
    )
    tx_qs = _transactions_scope(request.user).order_by("-created_at")

    if status_filter:
        orders_qs = orders_qs.filter(status=status_filter)
    if type_filter == "orders":
        tx_qs = tx_qs.none()
    elif type_filter == "transactions":
        orders_qs = orders_qs.none()

    if customer_filter and is_manager(request.user):
        balance_user = get_object_or_404(User, pk=customer_filter)
        orders_qs = orders_qs.filter(customer=balance_user)
        tx_qs = tx_qs.filter(customer=balance_user)

    customers_filter_list = (
        User.objects.filter(orders__isnull=False).select_related("customerprofile").distinct()
        if is_manager(request.user) else []
    )

    if negative_only and is_manager(request.user):
        negative_customer_ids = []
        for u in customers_filter_list:
            if compute_balance(u) < 0:
                negative_customer_ids.append(u.id)
        if negative_customer_ids:
            orders_qs = orders_qs.filter(customer_id__in=negative_customer_ids)
            tx_qs = tx_qs.filter(customer_id__in=negative_customer_ids)
        else:
            orders_qs = orders_qs.none()
            tx_qs = tx_qs.none()

    current_rate = get_current_eur_rate()
    orders_qs = _set_order_totals_uah(orders_qs, current_rate)
    filtered_balance = _transactions_total_uah(tx_qs) - _orders_total_uah_base(orders_qs, current_rate)
    proposal_tokens = {o.id: _proposal_token(o) for o in orders_qs}
    proposal_page_urls = {oid: reverse("orders:proposal_page", args=[tok]) for oid, tok in proposal_tokens.items()}
    proposal_excel_urls = {oid: reverse("orders:proposal_excel", args=[tok]) for oid, tok in proposal_tokens.items()}

    events = []
    for o in orders_qs:
        events.append({"type": "order", "created_at": o.created_at, "object": o})
    for tx in tx_qs:
        events.append({"type": "transaction", "created_at": tx.created_at, "object": tx})
    events.sort(key=lambda x: x["created_at"], reverse=True)

    balance_page_url = None
    if is_manager(request.user) and customer_filter:
        try:
            tok = _balance_token(balance_user.id)
            balance_page_url = request.build_absolute_uri(reverse("orders:balance_public", args=[tok]))
        except Exception:
            balance_page_url = None

    context = {
        "events": events,
        "statuses": Order.STATUS_CHOICES,
        "status_filter": status_filter,
        "customer_filter": customer_filter,
        "type_filter": type_filter,
        "negative_only": negative_only,
        "customer_options": customers_filter_list,
        "balance": compute_balance(balance_user),
        "balance_user": balance_user,
        # Ensure header balance reflects filtered customer (for managers)
        "user_balance": compute_balance(balance_user),
        "filtered_balance": filtered_balance,
        "status_badges": STATUS_BADGES,
        "status_labels": STATUS_LABELS,
        "proposal_page_urls": proposal_page_urls,
        "proposal_excel_urls": proposal_excel_urls,
        "balance_page_url": balance_page_url,
    }
    return render(request, "orders/balances_history.html", context)


@login_required
def balances_excel(request):
    """
    Export balances view (orders + transactions) to XLSX with current filters.
    """
    User = get_user_model()
    status_filter = request.GET.get("status") or ""
    customer_filter = request.GET.get("customer") or ""
    type_filter = request.GET.get("type") or ""
    negative_only = request.GET.get("negative") in ("1", "true", "on")
    balance_user = request.user

    orders_qs = (
        _orders_scope(request.user)
        .select_related("customer", "customer__customerprofile")
        .prefetch_related("status_logs", "component_items")
        .exclude(status=Order.STATUS_QUOTE)
        .order_by("-created_at")
    )
    tx_qs = _transactions_scope(request.user).order_by("-created_at")

    if status_filter:
        orders_qs = orders_qs.filter(status=status_filter)
    if type_filter == "orders":
        tx_qs = tx_qs.none()
    elif type_filter == "transactions":
        orders_qs = orders_qs.none()

    if customer_filter and is_manager(request.user):
        balance_user = get_object_or_404(User, pk=customer_filter)
        orders_qs = orders_qs.filter(customer=balance_user)
        tx_qs = tx_qs.filter(customer=balance_user)

    customers_filter_list = (
        User.objects.filter(orders__isnull=False).select_related("customerprofile").distinct()
        if is_manager(request.user) else []
    )

    if negative_only and is_manager(request.user):
        negative_customer_ids = []
        for u in customers_filter_list:
            if compute_balance(u) < 0:
                negative_customer_ids.append(u.id)
        if negative_customer_ids:
            orders_qs = orders_qs.filter(customer_id__in=negative_customer_ids)
            tx_qs = tx_qs.filter(customer_id__in=negative_customer_ids)
        else:
            orders_qs = orders_qs.none()
            tx_qs = tx_qs.none()

    current_rate = get_current_eur_rate()
    orders_qs = _set_order_totals_uah(orders_qs, current_rate)
    filtered_balance = _transactions_total_uah(tx_qs) - _orders_total_uah_base(orders_qs, current_rate)

    events = []
    for o in orders_qs:
        events.append({"type": "order", "created_at": o.created_at, "object": o})
    for tx in tx_qs:
        events.append({"type": "transaction", "created_at": tx.created_at, "object": tx})
    events.sort(key=lambda x: x["created_at"], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Баланс"

    headers = ["Дата", "Тип", "Опис", "Сума, грн"]
    ws.append(headers)
    bold = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = bold

    for e in events:
        if e["type"] == "order":
            o = e["object"]
            amount = -_round_uah_total(_order_base_total(o) * _order_rate(o, current_rate))
            desc = f"Замовлення №{o.id} ({o.get_status_display()})"
            type_label = "Замовлення"
        else:
            tx = e["object"]
            amount = _tx_amount_uah(tx)
            desc = f"Транзакція №{tx.id} ({tx.get_type_display()})"
            type_label = "Транзакція"
        created_at = e["created_at"]
        if hasattr(created_at, "tzinfo") and created_at.tzinfo:
            created_at = timezone.localtime(created_at).replace(tzinfo=None)
        ws.append([created_at, type_label, desc, float(amount)])

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="Разом").font = bold
    ws.cell(row=total_row, column=4, value=float(filtered_balance)).font = bold

    for col_letter in ["A", "B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Build filename with date and filters
    stamp = timezone.localtime(timezone.now()).strftime("%Y-%m-%d")
    name_parts = ["balance", stamp]
    if customer_filter:
        name_parts.append(f"customer-{customer_filter}")
    if status_filter:
        name_parts.append(f"status-{status_filter}")
    if type_filter:
        name_parts.append(f"type-{type_filter}")
    if negative_only:
        name_parts.append("negative")
    filename = "_".join(name_parts) + ".xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _balance_events_for_customer(customer, include_orders=True, include_transactions=True):
    orders_qs = (
        _orders_scope(customer)
        .filter(customer=customer)
        .select_related("customer", "customer__customerprofile")
        .prefetch_related("status_logs", "component_items")
        .exclude(status=Order.STATUS_QUOTE)
        .order_by("-created_at")
    )
    tx_qs = _transactions_scope(customer).filter(customer=customer).order_by("-created_at")
    if not include_orders:
        orders_qs = orders_qs.none()
    if not include_transactions:
        tx_qs = tx_qs.none()

    current_rate = get_current_eur_rate()
    orders_qs = _set_order_totals_uah(orders_qs, current_rate)
    events = []
    for o in orders_qs:
        created = o.created_at
        if hasattr(created, "tzinfo") and created.tzinfo:
            created = timezone.localtime(created).replace(tzinfo=None)
        amount = -_round_uah_total(_order_base_total(o) * _order_rate(o, current_rate))
        events.append({
            "type": "order",
            "created_at": created,
            "object": o,
            "description": f"Замовлення №{o.id} ({o.get_status_display()})",
            "amount_uah": amount,
            "amount_abs": abs(amount),
        })
    for tx in tx_qs:
        created = tx.created_at
        if hasattr(created, "tzinfo") and created.tzinfo:
            created = timezone.localtime(created).replace(tzinfo=None)
        amount = _tx_amount_uah(tx)
        events.append({
            "type": "transaction",
            "created_at": created,
            "object": tx,
            "description": f"Транзакція №{tx.id} ({tx.get_type_display()})",
            "amount_uah": amount,
            "amount_abs": abs(amount),
        })
    events.sort(key=lambda x: x["created_at"] or timezone.now(), reverse=True)

    balance_total = _transactions_total_uah(tx_qs) - _orders_total_uah_base(orders_qs, current_rate)
    return events, balance_total, current_rate


def balance_public_page(request, token: str):
    customer = _customer_from_balance_token(token)
    events, balance_total, current_rate = _balance_events_for_customer(customer)

    context = {
        "customer": customer,
        "profile": getattr(customer, "customerprofile", None),
        "events": events,
        "balance_total": balance_total,
        "current_rate": current_rate,
    }
    return render(request, "orders/balance_public.html", context)


@login_required
def balances_users(request):
    """
    EN: List of customers with balances and quick links to their transactions.
    UA: Список клієнтів із балансами та швидкими діями.
    """
    if not is_manager(request.user):
        messages.error(request, "Доступ заборонено.")
        return redirect("orders:balances")

    User = get_user_model()
    q = (request.GET.get("q") or "").strip()
    sort = request.GET.get("sort") or "-balance"

    users_qs = User.objects.filter(is_customer=True).select_related("customerprofile")
    if q:
        users_qs = users_qs.filter(
            Q(email__icontains=q)
            | Q(customerprofile__full_name__icontains=q)
            | Q(customerprofile__phone__icontains=q)
            | Q(customerprofile__company_name__icontains=q)
        )

    clients = list(users_qs)
    balances = {u.id: compute_balance(u, force_personal=True) for u in clients}

    def balance_val(user):
        return balances.get(user.id) or Decimal("0")

    sort_key = sort.lstrip("-")
    reverse = sort.startswith("-")

    if sort_key == "balance":
        clients.sort(key=lambda u: balance_val(u), reverse=reverse)
    elif sort_key in ("name", "full_name"):
        clients.sort(
            key=lambda u: (getattr(getattr(u, "customerprofile", None), "full_name", "") or u.email or "").lower(),
            reverse=reverse,
        )
    else:
        clients.sort(key=lambda u: (u.email or "").lower(), reverse=reverse)

    context = {
        "clients": clients,
        "balances": balances,
        "q": q,
        "sort": sort,
        "balance_tokens": {u.id: _balance_token(u.id) for u in clients},
    }
    return render(request, "orders/balances_users.html", context)



@login_required
def order_components_builder(request, pk):
    """
    EN: Builder for order components (sheet 'Комплектація').
    UA: Білдер комплектуючих для замовлення (аркуш 'Комплектація').
    """
    order = get_object_or_404(Order, pk=pk, deleted=False)
    readonly = bool(order and (not is_manager(request.user) and order.status != Order.STATUS_QUOTE))

    if request.method == "POST":
        if readonly:
            messages.warning(request, "Це замовлення можна лише переглядати.")
            return redirect("orders:order_components_builder", pk=order.pk)

        action = request.POST.get("status_action") or "save"
        chosen_customer = None
        can_pick_customer = is_manager(request.user) or request.user.is_staff or request.user.is_superuser
        if can_pick_customer:
            cust_id = request.POST.get("customer_id")
            if cust_id:
                try:
                    chosen_customer = get_user_model().objects.get(pk=cust_id)
                except get_user_model().DoesNotExist:
                    chosen_customer = None
            if chosen_customer is None:
                messages.error(request, "Оберіть клієнта для замовлення.")
                return redirect("orders:order_components_builder", pk=order.pk)
        components = parse_components_from_post(request.POST)
        base_note = (request.POST.get("note") or "").strip()
        if chosen_customer and can_pick_customer:
            base_note = re.sub(r"\s*\(створено менеджером [^)]+\)\s*$", "", base_note).strip()
            suffix = f" (створено менеджером {request.user.email})"
            base_note = (base_note + suffix).strip()
        order.note = base_note
        order.extra_service_label = (request.POST.get("extra_service_label") or "").strip()
        order.extra_service_amount_uah = _to_decimal(request.POST.get("extra_service_amount_uah"), default="0").quantize(Decimal("0.01"))
        markup_percent = _to_decimal(request.POST.get("markup_percent"), default=str(order.markup_percent or "0"))
        if chosen_customer and can_pick_customer:
            order.customer = chosen_customer
        target_customer = chosen_customer or order.customer
        if not chosen_customer:
            discount_pct_val = _normalize_discount_percent(getattr(order, "discount_percent", Decimal("0")))
        else:
            _, discount_pct_val = _customer_discount_multiplier(target_customer)
        discount_multiplier, _ = _customer_discount_multiplier(pct=discount_pct_val)

        # EN: Replace all existing components with new list
        # UA: Повністю замінюємо поточний список комплектуючих новим
        OrderComponentItem.objects.filter(order=order).delete()

        total_eur = Decimal("0")
        bulk = []
        for row in components:
            price_raw = Decimal(row["price_eur"] or 0)
            qty = Decimal(row["quantity"] or 0)
            price_eur = (price_raw * discount_multiplier).quantize(Decimal("0.01"))
            total_eur += price_eur * qty
            bulk.append(
                OrderComponentItem(
                    order=order,
                    name=row["name"],
                    color=row["color"],
                    unit=row["unit"],
                    price_eur=price_eur,
                    quantity=qty,
                )
            )

        if bulk:
            OrderComponentItem.objects.bulk_create(bulk)
        # Save total for order (used in listings/balance)
        order.total_eur = total_eur.quantize(Decimal("0.01"))
        order.eur_rate = order.eur_rate or get_current_eur_rate()
        order.markup_percent = markup_percent.quantize(Decimal("0.01"))

        # Status transitions (mirror roller orders)
        new_status = None
        if action == "to_work":
            new_status = Order.STATUS_IN_WORK
        elif action == "next":
            new_status = order.next_status()
        elif action == "prev":
            new_status = order.prev_status()
        # Default save keeps current status

        update_fields = ["total_eur", "eur_rate", "markup_percent", "note", "extra_service_label", "extra_service_amount_uah", "discount_percent"]
        if chosen_customer and can_pick_customer:
            update_fields.append("customer")
        order.discount_percent = discount_pct_val
        order.save(update_fields=update_fields)

        if new_status and not _apply_status_change(order, new_status, request):
            return redirect(reverse("orders:order_components_builder", kwargs={"pk": order.pk}))

        messages.success(request, "Замовлення збережено.")

        # EN: Redirect back to builder or to order detail — adjust as needed
        # UA: Редірект назад у білдер або на сторінку замовлення — за потреби зміни
        next_url = (request.POST.get("next_url") or "").strip()
        if next_url and url_has_allowed_host_and_scheme(
            next_url,
            allowed_hosts={request.get_host()},
            require_https=request.is_secure(),
        ):
            return redirect(next_url)
        return redirect(reverse("orders:order_components_builder", kwargs={"pk": order.pk}))

    # ---------- GET: формируем components_json для фронта ----------

    # EN: Load existing components for edit mode
    # UA: Завантажуємо існуючі комплектуючі (режим редагування)
    items = (
      order.component_items.all()
      .order_by("id")
    )

    components_payload = []
    for item in items:
        components_payload.append(
            {
                "name": item.name,
                "color": item.color,
                "unit": item.unit,
                "price_eur": str(item.price_eur),
                "quantity": str(item.quantity),
            }
        )

    payment_prompt = request.session.pop("payment_prompt", None)
    shortage_ctx = _payment_shortage_context(order)
    proposal_token = _proposal_token(order)
    proposal_page_url = reverse("orders:proposal_page", args=[proposal_token])
    proposal_excel_url = reverse("orders:proposal_excel", args=[proposal_token])
    context = {
        "order": order,
        "PRICE_SHEET_URL": PRICE_SHEET_URL,
        "components_json": json.dumps(components_payload, ensure_ascii=False),
        "status_logs": order.status_logs.all(),
        "readonly": readonly,
        "next_status_label": STATUS_LABELS.get(order.next_status()) if order else None,
        "builder_rate": order.eur_rate or get_current_eur_rate(),
        "payment_message_text": (payment_prompt or {}).get("text") or _get_payment_message_text(),
        "payment_shortage": payment_prompt or shortage_ctx,
        "proposal_page_url": proposal_page_url,
        "proposal_excel_url": proposal_excel_url,
        "customers_filter_list": list(
            CustomerProfile.objects.select_related("user").order_by("full_name", "user__email")
        ) if is_manager(request.user) else [],
    }

    return render(request, "orders/components_builder.html", context)


@login_required
def update_status_preview(request, pk):
    """
    Quick status update to 'in_work' from preview without opening builder.
    """
    redirect_back = request.META.get("HTTP_REFERER") or reverse("orders:list")
    if request.method not in ("POST", "GET"):
        messages.warning(request, "Невірний метод запиту.")
        return redirect(redirect_back)

    if is_manager(request.user):
        order = get_object_or_404(Order, pk=pk, deleted=False)
    else:
        order = get_object_or_404(Order, pk=pk, customer=request.user, deleted=False)

    if request.method == "GET" and "status_action" not in request.GET:
        messages.warning(request, "Невірний запит.")
        return redirect(redirect_back)

    payload = request.POST if request.method == "POST" else request.GET
    action = (payload.get("status_action") or "next").strip()
    show_payment_prompt = (payload.get("show_payment_prompt") or "").strip()

    # Клиент может только перевести прорахунок у роботу
    if not is_manager(request.user) and action != "to_work":
        messages.warning(request, "Дія недоступна.")
        return redirect(redirect_back)

    new_status = None
    if action == "to_work":
        new_status = Order.STATUS_IN_WORK
    elif action == "prev":
        new_status = order.prev_status()
    else:  # next (default)
        new_status = order.next_status()

    if not new_status:
        messages.warning(request, "Не можна змінити статус.")
        return redirect(redirect_back)

    if not _apply_status_change(order, new_status, request):
        if show_payment_prompt == "1" and action == "to_work":
            if is_manager(request.user):
                return redirect(redirect_back)
            shortage_ctx = _payment_shortage_context(order)
            if shortage_ctx:
                request.session["payment_prompt"] = {
                    "text": _get_payment_message_text() or "",
                    "shortage": str(shortage_ctx.get("shortage") or ""),
                    "orders": shortage_ctx.get("orders") or [],
                }
                if order.component_items.exists():
                    return redirect("orders:order_components_builder", pk=order.pk)
                return redirect("orders:builder_edit", pk=order.pk)
        return redirect(redirect_back)

    messages.success(request, f"Замовлення №{order.pk} переведено в роботу.")
    return redirect(redirect_back)


@login_required
def update_status_bulk(request):
    """
    EN: Bulk move selected orders to "in_work" from list.
    UA: Масове переведення вибраних замовлень у "В роботі" зі списку.
    """
    if request.method != "POST":
        messages.warning(request, "Невірний метод запиту.")
        return redirect(reverse("orders:list"))

    order_ids = request.POST.getlist("order_ids")
    list_mode = request.POST.get("list_mode") or ""
    fallback = "orders:components_list" if list_mode == "components" else "orders:list"
    next_url = (request.POST.get("next_url") or "").strip()
    if not next_url or not url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        next_url = reverse(fallback)

    if not order_ids:
        messages.warning(request, "Оберіть хоча б одне замовлення.")
        return redirect(next_url)

    qs = Order.objects.filter(pk__in=order_ids, deleted=False)
    if not is_manager(request.user):
        qs = qs.filter(customer=request.user)

    success = 0
    skipped = 0
    failed = 0

    for order in qs:
        if order.status != Order.STATUS_QUOTE:
            skipped += 1
            continue
        if _apply_status_change(order, Order.STATUS_IN_WORK, request):
            success += 1
        else:
            failed += 1

    if success:
        messages.success(request, f"Запущено в роботу: {success}.")
    if skipped:
        messages.info(request, f"Пропущено (не в прорахунку): {skipped}.")
    if failed:
        messages.warning(request, f"Не вдалося запустити: {failed}.")

    return redirect(next_url)


def order_proposal_page(request, token: str):
    """
    Public read-only commercial proposal page (no auth required).
    """
    order = _order_from_token(token)
    profile = getattr(order.customer, "customerprofile", None)
    customer_name = getattr(profile, "company_name", "") or getattr(profile, "full_name", "") or str(order.customer)
    customer_phone = getattr(profile, "phone", "") or ""
    customer_full_name = getattr(profile, "full_name", "") or str(order.customer)
    customer_address = getattr(profile, "trade_address", "") or ""
    rate = _order_rate(order, get_current_eur_rate())
    base_total_eur = _order_base_total(order)
    markup = Decimal(order.markup_percent or 0)
    markup_multiplier = Decimal("1") + (markup / Decimal("100"))
    rate_with_markup = Decimal(rate or 0) * markup_multiplier
    total_with_markup_eur = base_total_eur * (Decimal("1") + markup / Decimal("100"))
    extra_uah = Decimal(getattr(order, "extra_service_amount_uah", 0) or 0)
    total_base_uah = _round_uah_total(base_total_eur * rate + extra_uah)
    total_with_markup_uah = _round_uah_total(total_with_markup_eur * rate + extra_uah)

    items_payload = []
    for idx, it in enumerate(order.items.all(), start=1):
        options = _collect_item_options(it, rate)
        subtotal_eur = Decimal(it.subtotal_eur or 0)
        sys_lower = (it.system_sheet or "").lower()
        is_flat_system = "плоска" in sys_lower
        width_label_suffix = ""
        height_label_suffix = ""
        if is_flat_system and it.gabarit_width_flag:
            width_label_suffix = " (по тканині)"
        elif (not is_flat_system) and it.gabarit_width_flag:
            width_label_suffix = " (габаритна)"
        if is_flat_system and it.fabric_height_flag:
            height_label_suffix = " (по тканині)"
        items_payload.append(
            {
                "index": idx,
                "system_sheet": it.system_sheet,
                "table_section": it.table_section,
                "fabric_name": it.fabric_name,
                "fabric_color_code": it.fabric_color_code,
                "width_fabric_mm": it.width_fabric_mm,
                "width_label_suffix": width_label_suffix,
                "height_gabarit_mm": it.height_gabarit_mm,
                "height_label_suffix": height_label_suffix,
                "gabarit_width_flag": it.gabarit_width_flag,
                "fabric_height_flag": it.fabric_height_flag,
                "roll_height_info": it.roll_height_info,
                "control_side": it.control_side,
                "control_side_label": _control_side_label(it.control_side),
                "quantity": it.quantity,
                "subtotal_eur": subtotal_eur,
                "subtotal_uah": _round_uah_total(subtotal_eur * rate_with_markup),
                "note": it.note,
                "options": _collect_item_options(it, rate, markup_multiplier),
            }
        )

    components_payload = []
    for comp in order.component_items.all():
        price_eur = Decimal(comp.price_eur or 0)
        qty = Decimal(comp.quantity or 0)
        total_eur = price_eur * qty
        components_payload.append(
            {
                "name": comp.name,
                "color": comp.color,
                "unit": comp.unit,
                "price_eur": price_eur,
                "quantity": qty,
                "total_eur": total_eur,
                "total_uah": _round_uah_total(total_eur * rate_with_markup),
            }
        )

    proposal_excel_url = reverse("orders:proposal_excel", args=[token])
    context = {
        "order": order,
        "customer_info": {
            "name": customer_name,
            "phone": customer_phone,
            "full_name": customer_full_name,
            "address": customer_address,
        },
        "rate": rate,
        "base_total_eur": base_total_eur,
        "total_with_markup_eur": total_with_markup_eur,
        "total_base_uah": total_base_uah,
        "total_with_markup_uah": total_with_markup_uah,
        "markup_percent": markup,
        "rate_with_markup": rate_with_markup,
        "items": items_payload,
        "components": components_payload,
        "proposal_excel_url": proposal_excel_url,
    }
    return render(request, "orders/proposal_public.html", context)


def order_proposal_excel(request, token: str):
    """Public XLSX download for commercial proposal."""
    order = _order_from_token(token)
    filename, data = _build_proposal_workbook(order)
    response = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    return response


@login_required
def order_components_builder_new(request):
    """
    EN: Create new order and go to components builder.
    UA: Створює нове замовлення і переходить у білдер комплектуючих.
    """
    chosen_customer = request.user
    can_pick_customer = is_manager(request.user) or request.user.is_staff or request.user.is_superuser
    if can_pick_customer:
        cust_id = request.GET.get("customer") or request.POST.get("customer_id")
        if cust_id:
            try:
                chosen_customer = get_user_model().objects.get(pk=cust_id)
            except get_user_model().DoesNotExist:
                chosen_customer = request.user

    _, discount_pct_val = _customer_discount_multiplier(chosen_customer)

    order = Order.objects.create(
        customer=chosen_customer,        # ✔ обязателен, иначе IntegrityError по customer_id
        title="Замовлення (комплектуючі)",  # любой не пустой заголовок
        description="",
        status=Order.STATUS_QUOTE,
        eur_rate_at_creation=get_current_eur_rate(),
        eur_rate=get_current_eur_rate(),
        markup_percent=Decimal("0"),
        total_eur=Decimal("0"),
        discount_percent=discount_pct_val,
    )
    return redirect("orders:order_components_builder", pk=order.pk)


@staff_member_required  # EN: only staff can update; UA: тільки staff-користувачі можуть оновлювати
@require_POST
def update_eur_rate_view(request):
    """
    EN: Update EUR rate from NBU and redirect back.
    UA: Оновлює курс EUR з НБУ та повертає назад.
    """
    mode = request.POST.get("mode") or "online"
    if mode == "manual":
        rate = request.POST.get("rate")
        try:
            rate_val = Decimal(str(rate).replace(",", "."))
            if rate_val <= 0:
                raise InvalidOperation
            obj, _ = CurrencyRate.objects.update_or_create(
                currency="EUR",
                defaults={"rate_uah": rate_val, "source": "manual"},
            )
            messages.success(
                request,
                f"Курс EUR збережено вручну: {obj.rate_uah} грн",
            )
        except Exception:
            messages.error(request, "Введіть коректний курс EUR (приклад: 39.50)")
    else:
        try:
            obj = update_eur_rate_from_nbu()
            messages.success(
                request,
                f"Курс EUR оновлено: {obj.rate_uah} грн (джерело: {obj.source})"
            )
        except Exception as e:
            messages.error(
                request,
                f"Не вдалося оновити курс EUR: {e}"
            )
    # Log history if we have an object and no errors were raised
    if "obj" in locals():
        CurrencyRateHistory.objects.create(
            currency=obj.currency,
            rate_uah=obj.rate_uah,
            mode=mode,
            source=obj.source,
            user=request.user if request.user.is_authenticated else None,
        )

    # EN: redirect back where user came from; UA: повертаємось туди, звідки прийшли
    next_url = request.META.get("HTTP_REFERER") or reverse("core:dashboard")
    return redirect(next_url)


@login_required
def currency_rate_history(request):
    if is_manager(request.user):
        history = CurrencyRateHistory.objects.select_related("user").all()
        show_rate_history = True
    else:
        history = CurrencyRateHistory.objects.none()
        show_rate_history = False
    return render(
        request,
        "orders/currency_history.html",
        {"history": history, "show_rate_history": show_rate_history},
    )
