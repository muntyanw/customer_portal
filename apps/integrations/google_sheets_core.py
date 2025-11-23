# -*- coding: utf-8 -*-
# EN: Google Sheets core helpers (download XLSX, cache, list sheet titles, low-level cell helpers)
# UA: Базові хелпери Google Sheets (завантаження XLSX, кеш, список вкладок, низькорівневі хелпери)

from __future__ import annotations

import io
import os
import re
import json
import time
import hashlib
import logging
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import List, Optional

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from django.conf import settings

logger = logging.getLogger("sheets")

Q = Decimal
CACHE_TTL_SECONDS = 300  # 5 min; can be changed

CACHE_DIR = getattr(
    settings,
    "SHEETS_CACHE_DIR",
    os.path.join(settings.BASE_DIR, "tmp", "sheets_cache"),
)
os.makedirs(CACHE_DIR, exist_ok=True)


def round_money(x: Decimal) -> Decimal:
    """
    EN: Round money to 0.01 with HALF_UP.
    UA: Округлення грошей до 0.01 з режимом HALF_UP.
    """
    return x.quantize(Q("0.01"), rounding=ROUND_HALF_UP)


def _xlsx_export_url(google_sheet_url: str) -> str:
    """
    EN: Build XLSX export URL from public Google Sheets URL.
    UA: Формує посилання експорту XLSX з публічного URL Google Sheets.
    """
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", google_sheet_url)
    if not m:
        raise ValueError("Bad Google Sheets URL")
    sheet_id = m.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def _cache_keys(google_sheet_url: str):
    """
    EN: Return paths for XLSX and META cache files.
    UA: Повертає шляхи до файлів кешу XLSX та META.
    """
    key = hashlib.sha256(google_sheet_url.encode("utf-8")).hexdigest()
    xlsx_path = os.path.join(CACHE_DIR, f"{key}.xlsx")
    meta_path = os.path.join(CACHE_DIR, f"{key}.json")
    return xlsx_path, meta_path


def _read_meta(meta_path: str) -> dict:
    """
    EN: Read JSON metadata for cache (ETag, ts).
    UA: Читає JSON-метадані кешу (ETag, ts).
    """
    if not os.path.exists(meta_path):
        return {}
    try:
        with open(meta_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logger.error("_read_meta failed: %s", e, exc_info=True)
        return {}


def _write_meta(meta_path: str, data: dict):
    """
    EN: Atomic write of cache metadata.
    UA: Атомарний запис метаданих кешу.
    """
    tmp = meta_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f)
    os.replace(tmp, meta_path)


def _download_workbook(google_sheet_url: str, force_refresh: bool = False):
    """
    EN: Download XLSX with ETag/TTL cache. If Google Sheets is unavailable,
        fallback to the last cached local file.
    UA: Завантажує XLSX з кешем через ETag/TTL. Якщо Google Sheets недоступний,
        відкриває останній локальний файл.
    """
    url = _xlsx_export_url(google_sheet_url)
    xlsx_path, meta_path = _cache_keys(google_sheet_url)
    meta = _read_meta(meta_path)
    etag = meta.get("etag")
    ts = meta.get("ts", 0)

    # If TTL valid and file exists – use cached version
    if (
        not force_refresh
        and os.path.exists(xlsx_path)
        and time.time() - ts < CACHE_TTL_SECONDS
    ):
        try:
            with open(xlsx_path, "rb") as f:
                return load_workbook(io.BytesIO(f.read()), data_only=True)
        except Exception:
            pass  # fallback to trying download

    # --- Try downloading Google Sheet ---
    headers = {}
    if etag and not force_refresh:
        headers["If-None-Match"] = etag

    try:
        resp = requests.get(url, timeout=30, headers=headers)
        # 304 Not Modified → update TTL and load cached
        if resp.status_code == 304 and os.path.exists(xlsx_path):
            meta["ts"] = time.time()
            _write_meta(meta_path, meta)
            with open(xlsx_path, "rb") as f:
                return load_workbook(io.BytesIO(f.read()), data_only=True)

        resp.raise_for_status()

        # Successfully downloaded → save and update cache
        content = resp.content
        with open(xlsx_path, "wb") as f:
            f.write(content)

        new_etag = resp.headers.get("ETag")
        meta = {"etag": new_etag, "ts": time.time()}
        _write_meta(meta_path, meta)

        return load_workbook(io.BytesIO(content), data_only=True)

    except Exception as ex:
        # --- Fallback when offline or Google unreachable ---
        if os.path.exists(xlsx_path):
            # Log warning but continue working offline
            logger.warning(
                f"Google Sheets fetch failed ({ex}); using cached file: {xlsx_path}"
            )
            try:
                with open(xlsx_path, "rb") as f:
                    return load_workbook(io.BytesIO(f.read()), data_only=True)
            except Exception as ex2:
                raise RuntimeError(
                    f"Local cache exists but cannot be opened: {xlsx_path}"
                ) from ex2

        # No internet + no cache
        raise RuntimeError(
            "Failed to download Google Sheet and no local cache available."
        ) from ex

def list_sheet_titles(google_sheet_url: str, *, force_refresh: bool = False) -> List[str]:
    """
    EN: Return list of sheet names (tabs).
    UA: Повертає список назв вкладок.
    """
    wb = _download_workbook(google_sheet_url, force_refresh=force_refresh)
    return wb.sheetnames


def _row_values(
    ws: Worksheet,
    row: int,
    start_col: int = 1,
    end_col: Optional[int] = None,
) -> List[Optional[str]]:
    """
    EN: Raw row read helper (no business logic).
    UA: Низькорівневий хелпер читання рядка (без бізнес-логіки).
    """
    if end_col is None:
        end_col = ws.max_column
    vals: List[Optional[str]] = []
    for c in range(start_col, end_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is None or isinstance(v, str):
            vals.append(v)
        else:
            vals.append(str(v))
    return vals


def _to_decimal(x) -> Optional[Decimal]:
    """
    EN: Convert any cell-like value to Decimal (generic).
    UA: Перетворення будь-якого значення комірки у Decimal (загальний хелпер).
    """
    if x is None:
        return None

    s = str(x).strip()
    if not s:
        return None

    # normalize decimal separator
    s = s.replace(",", ".")
    # remove spaces and NBSPs
    s = re.sub(r"[\s\u00A0\u202F]", "", s)
    # keep digits, dot, minus
    s = re.sub(r"[^0-9\.\-]", "", s)

    if s in ("", ".", "-", "-.", ".-"):
        return None

    try:
        return Decimal(s)
    except InvalidOperation:
        logger.error("_to_decimal failed to parse cleaned value: %r", s)
        return None
    
    
def get_money_value(ws: Worksheet, row: int, col: int) -> Optional[Decimal]:
    """
    EN: Read a cell and return it as Decimal money (rounded to 0.01).
    UA: Читає комірку і повертає її значення як Decimal (гроші, округлені до 0.01).
    """
    raw = ws.cell(row=row, column=col).value
    dec = _to_decimal(raw)
    if dec is None:
        return None
    return round_money(dec)


def col_letter_to_index(col: str) -> int:
    """
    EN: Convert Excel column letter(s) to a 1-based column index.
    UA: Перетворює літерний індекс стовпця Excel у числовий (з 1).
    """
    if not col or not isinstance(col, str):
        raise ValueError("Column must be a non-empty string")

    col = col.strip().upper()
    result = 0

    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column letter: {ch}")
        result = result * 26 + (ord(ch) - ord("A") + 1)

    return result


def get_str_value(ws: Worksheet, row: int, col: int) -> Optional[str]:
    """
    EN: Read a cell and return its value as a cleaned string (strip whitespace).
    UA: Зчитує комірку і повертає значення як рядок (обрізаючи пробіли).
    """
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    return str(v).strip()


def get_str_values(
    ws: Worksheet,
    row_start: int,
    row_end: int,
    col_start: int,
    col_end: int
) -> str:
    """
    EN: Read a 2D cell range and return rows joined by <br/>.
    UA: Зчитує 2D-діапазон і повертає рядки, з’єднані через <br/>.
    """
    rows_out = []

    for r in range(row_start, row_end + 1):
        parts = []
        for c in range(col_start, col_end + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                parts.append("")
            else:
                parts.append(str(v).strip())
        # объединяем значения одной строки пробелом или просто склеиваем
        row_text = " ".join(p for p in parts if p)
        rows_out.append(row_text)

    return "<br/>".join(rows_out)





    
    

