# -*- coding: utf-8 -*-
# EN: Google Sheets price parser (sections, fabrics, price preview)
# UA: Парсер прайсу Google Sheets (секції, тканини, прев'ю ціни)

from __future__ import annotations

import re
from typing import List, Dict, Optional, Any


from openpyxl.worksheet.worksheet import Worksheet

from apps.sheet_config import sheetName, sheetConfigs, getConfigBySheetName, sheetConfig
from .google_sheets_core import (
    _download_workbook,
    _row_values,
    _to_decimal,
    round_money,
    Q,
    logger,
    col_letter_to_index,
    get_money_value,
    get_str_values,
)


def _norm_title(s: Optional[str]) -> str:
    """
    EN: Normalize title for comparison.
    UA: Нормалізує заголовок для порівняння.
    """
    return (s or "").strip().lower()


# ========= HELPERS / SECTION FINDERS =========


def find_sections_merged(
    ws: Worksheet,
    *,
    min_merged_width: int = 12,
) -> List[Dict]:
    """
    UA: Знаходить секції лише за merged-заголовками в одному рядку.
    EN: Find sections only by one-line merged headers (width >= min_merged_width).
    Returns: [{"title": str, "row": int, "col": int}, ...]
    """
    sections: List[Dict] = []
    seen = set()

    for rng in ws.merged_cells.ranges:
        if rng.min_row == rng.max_row:
            width = rng.max_col - rng.min_col + 1
            if width >= min_merged_width:
                v = ws.cell(row=rng.min_row, column=rng.min_col).value
                if isinstance(v, str) and v.strip():
                    key = (rng.min_row, rng.min_col)
                    if key not in seen:
                        seen.add(key)
                        sections.append(
                            {
                                "title": v.strip(),
                                "row": rng.min_row,
                                "col": rng.min_col,
                            }
                        )

    sections.sort(key=lambda x: (x["row"], x["col"]))
    return sections


def find_sections_by_headers(
    ws: Worksheet,
    *,
    title_prefix: str = "",
    min_merged_width: int = 0,  # kept for compatibility, not used
    search_cols: int = 6,
    case_insensitive: bool = True,
    sheet_name: str = "",  # 🔹 новый параметр
) -> List[Dict]:
    """
    UA: Шукає секції за текстовими заголовками (без прив'язки до merged-ячейок).
        Використовується для пошуку рядків типу:
            "Фальш-ролети, біла система"
            "Фальш-ролети, коричнева система"
            ...
        Додатково: рядок вважається секцією, якщо перші 5 символів
        збігаються з першими 5 символами назви аркуша (sheet_name).
    EN: Find sections by textual headers (no dependency on merged cells).
        Additionally, a row is considered a section only if the first
        5 characters match the first 5 characters of sheet_name.
    Returns: [{"title": str, "row": int, "col": int}, ...]
    """
    sections: List[Dict] = []

    prefix_norm = _norm_title(title_prefix)

    # 🔹 Нормализуем sheet_name один раз
    sheet_prefix_norm = ""
    if sheet_name:
        sheet_norm_full = sheet_name.strip()
        if case_insensitive:
            sheet_norm_full = sheet_norm_full.lower()
        sheet_prefix_norm = sheet_norm_full[:4]

    for r in range(1, ws.max_row + 1):
        for c in range(1, search_cols + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            raw = v.strip()
            if not raw:
                continue

            norm = raw.lower() if case_insensitive else raw

            # 🔹 Условие по title_prefix, как было
            if prefix_norm and not norm.startswith(prefix_norm):
                continue

            # 🔹 Новое условие: первые 4 символов совпадают с sheet_name
            if sheet_prefix_norm:
                if norm[:4] != sheet_prefix_norm:
                    continue

            # Heuristic: must contain word "система"
            if "система" not in norm:
                continue

            sections.append(
                {
                    "title": raw,
                    "row": r,
                    "col": c,
                }
            )
            break

    sections.sort(key=lambda x: (x["row"], x["col"]))
    return sections


def pick_width_band(width_bands: List[str], width_mm: int) -> Optional[int]:
    """
    UA: Визначає індекс смуги ширини за текстом ('До 400мм', '401–450' ...).
    EN: Chooses band index from textual ranges.
    """
    for i, b in enumerate(width_bands):
        b = str(b).replace(" ", "").replace("мм", "")
        if b.startswith("До"):
            try:
                limit = int(re.sub(r"^До", "", b))
                if width_mm <= limit:
                    return i
            except Exception as e:
                logger.error("pick_width_band failed: %s", e, exc_info=True)
                pass
        else:
            m = re.match(r"(\d+)[–-](\d+)", b)
            if m:
                lo, hi = int(m.group(1)), int(m.group(2))
                if lo <= width_mm <= hi:
                    return i
    return None


def _width_out_of_range_error(gabarit_limit_mm: Optional[int], roll_height_mm: Optional[int]) -> ValueError:
    """
    UA: Формує повідомлення про вихід ширини за межі прайсу з контекстом висот із прайсу.
    EN: Build width-out-of-range error with extra context (from price table).
    """
    parts = []
    if gabarit_limit_mm:
        parts.append(f"Габаритна висота (прайс): {gabarit_limit_mm} мм")
    if roll_height_mm:
        parts.append(f"Висота рулону: {roll_height_mm} мм")
    extra = f" ({'; '.join(parts)})" if parts else ""
    return ValueError(f"Ширина поза діапазонами прайсу{extra}")


def _compute_price_detail(
    *,
    fabric: Dict,
    width_mm: int,
    gabarit_height_mm: int,
    bands: List[Any],
    band_idx: int,
    base_cell: Any,
    magnets_price: Optional[Q] = None,
) -> Dict[str, Any]:
    """
    UA: Розрахунок бази/доплати + перевірка ширини по висоті рулону.
    EN: Compute base/surcharge and enforce roll-height width limit.
    """
    if base_cell is None:
        raise ValueError("Ціна відсутня у вибраній смузі")

    limit = fabric.get("gabarit_limit_mm") or 0
    roll_mm = fabric.get("roll_height_mm") or 0
    base = Q(base_cell)

    if gabarit_height_mm <= limit:
        surcharge = Q("0.00")
    else:
        over = gabarit_height_mm - limit
        steps = (over + 99) // 100  # кожні 10см, заокруглення догори
        surcharge = round_money(base * Q("0.10") * Q(int(steps)))

    # Якщо висота виробу перевищує габаритну — ширина обмежується висотою рулону
    if gabarit_height_mm > limit and roll_mm and width_mm > roll_mm:
        raise ValueError(
            f"При перевищенні габаритної висоти ширина не може бути більшою за висоту рулону "
            f"(Габаритна висота (прайс): {limit} мм; Висота рулону: {roll_mm} мм)"
        )

    return {
        "gabarit_limit_mm": limit or None,
        "roll_height_mm": roll_mm or None,
        "band_index": band_idx,
        "band_label": bands[band_idx] if bands and band_idx < len(bands) else None,
        "base_price_eur": str(round_money(base)),
        "surcharge_height_eur": str(surcharge),
        "magnets_price_eur": str(round_money(magnets_price or Q("0.00"))),
    }


def fillOptions(sheet_name, result, ws, header_row, section_title=""):
    
        # 6) Extra magnets price (for Falshi sheet only)
    if sheet_name == sheetName.falshi:
        result["magnets_price_eur"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("D")
        )
        result["comment_system_red"] = get_str_values(
            ws, header_row - 3, header_row - 3, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 2, header_row - 2, 1, 1
        )

    if sheet_name == sheetName.falshiDn:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 3, header_row - 3, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 2, header_row - 2, 1, 1
        )

    if sheet_name == sheetName.vidkr19yiBesta:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 9, header_row - 9, 1, 1
        )
        
        result["comment_system_red"] += "<br/> Кліпса кріплення для верхньої планки ПВХ, Верхня планка ПВХ зі скотчем (монтаж без свердління) існує тільки в білому і коричневому кольорах!"
        
        result["comment_system_green"] = get_str_values(
            ws, header_row - 8, header_row - 6, 1, 1
        )
        result["cord_pvc_tension_price_eur"] = get_money_value(
            ws, header_row - 3, col_letter_to_index("D")
        )
        result["cord_copper_barrel_price_eur"] = get_money_value(
            ws, header_row - 2, col_letter_to_index("D")
        )
        result["magnets_price_eur"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("D")
        )
        
        result["top_pvc_clip_pair_price_eur"] = get_money_value(
            ws, header_row - 3, col_letter_to_index("N")
        )
        result["top_pvc_bar_tape_price_eur_mp"] = get_money_value(
            ws, header_row - 2, col_letter_to_index("N")
        )
        result["bottom_wide_bar_price_eur_mp"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("N")
        )

    if sheet_name == sheetName.vidkr19yiBestaDn:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 8, header_row - 7, 1, 1
        )
        result["comment_system_red"] += "<br/>" + get_str_values(
            ws,
            header_row - 4,
            header_row - 4,
            col_letter_to_index("E"),
            col_letter_to_index("E"),
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 7, header_row - 5, 1, 1
        )
        result["metal_cord_fix_price_eur"] = get_money_value(
            ws, header_row - 3, col_letter_to_index("D")
        )
        result["top_pvc_clip_pair_price_eur"] = get_money_value(
            ws, header_row - 2, col_letter_to_index("D")
        )
        result["top_bar_scotch_price_eur_mp"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("D")
        )

    if sheet_name == sheetName.zakrytaPloskaBesta:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 6, header_row - 6, 1, 1
        )
        result["comment_system_red"] += "<br/>" + get_str_values(
            ws, header_row - 4, header_row - 4, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 5, header_row - 5, 1, 1
        )
        result["comment_system_green"] += "<br/>" + get_str_values(
            ws, header_row - 3, header_row - 2, 1, 1
        )

    if sheet_name == sheetName.zakrytaPloskaBestaDn:
        if "біла" in section_title:
            result["comment_system_red"] = get_str_values(
                ws, header_row - 5, header_row - 5, 1, 1
            )
            result["comment_system_green"] = get_str_values(
                ws, header_row - 4, header_row - 2, 1, 1
            )
        else:
            result["comment_system_red"] = get_str_values(
                ws, header_row - 6, header_row - 6, 1, 1
            )
            result["comment_system_red"] += "<br/>" + get_str_values(
                ws, header_row - 4, header_row - 4, 1, 1
            )
            result["comment_system_green"] = get_str_values(
                ws, header_row - 5, header_row - 5, 1, 1
            )
            result["comment_system_green"] += "<br/>" + get_str_values(
                ws, header_row - 3, header_row - 2, 1, 1
            )

    if sheet_name == sheetName.zakrytaPpodibBesta:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 7, header_row - 7, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 6, header_row - 2, 1, 1
        )

    if sheet_name == sheetName.zakrytaPpodibnaBestaDn:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 6, header_row - 6, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 5, header_row - 2, 1, 1
        )

    if sheet_name == sheetName.vidkr25yiBesta:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 8, header_row - 8, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 7, header_row - 4, 1, 1
        )

        result["cord_pvc_tension_price_eur"] = get_money_value(
            ws, header_row - 3, col_letter_to_index("D")
        )
        result["cord_copper_barrel_price_eur"] = get_money_value(
            ws, header_row - 2, col_letter_to_index("D")
        )
        result["metal_kronsht_price_eur"] = get_money_value(
            ws, header_row - 2, col_letter_to_index("N")
        )
        result["magnets_price_eur"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("D")
        )
        
        result["bottom_wide_bar_price_eur_mp"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("N")
        )

    if sheet_name == sheetName.vidkr25yiDn:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 6, header_row - 6, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 5, header_row - 3, 1, 1
        )

        result["metal_cord_fix_price_eur"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("D")
        )
        
    if sheet_name == sheetName.vidkrPruzhynna:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 5, header_row - 5, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 4, header_row - 2, 1, 1
        )

    if sheet_name == sheetName.zakrPruzhPpodibBesta:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 7, header_row - 7, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 6, header_row - 2, 1, 1
        )
        
    if sheet_name == sheetName.vidkr32yiLouvolitte:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 8, header_row - 8, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 7, header_row - 5, 1, 1
        )

        result["cord_pvc_tension_price_eur"] = get_money_value(
            ws, header_row - 3, col_letter_to_index("D")
        )
        result["cord_copper_barrel_price_eur"] = get_money_value(
            ws, header_row - 2, col_letter_to_index("D")
        )
        result["magnets_price_eur"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("D")
        )
        result["bottom_wide_bar_price_eur_mp"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("N")
        )
        
    if sheet_name == sheetName.vidkr47yiDvyhunAboLouvolit:
        result["comment_system_red"] = get_str_values(
            ws, header_row - 8, header_row - 8, 1, 1
        )
        result["comment_system_green"] = get_str_values(
            ws, header_row - 7, header_row - 5, 1, 1
        )

        result["cord_pvc_tension_price_eur"] = get_money_value(
            ws, header_row - 3, col_letter_to_index("D")
        )
        result["cord_copper_barrel_price_eur"] = get_money_value(
            ws, header_row - 2, col_letter_to_index("D")
        )
        result["magnets_price_eur"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("D")
        )
        result["motor_no_remote_price_eur"] = get_money_value(
            ws, header_row - 3, col_letter_to_index("N")
        )
        result["motor_with_remote_price_eur"] = get_money_value(
            ws, header_row - 2, col_letter_to_index("N")
        )
        result["bottom_wide_bar_price_eur_mp"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("N")
        )
        result["remote_5ch_price_eur"] = get_money_value(
            ws, header_row - 3, col_letter_to_index("X")
        )
        result["remote_15ch_price_eur"] = get_money_value(
            ws, header_row - 2, col_letter_to_index("X")
        )
        result["middle_bracket_price_eur"] = get_money_value(
            ws, header_row - 1, col_letter_to_index("X")
        )
        
    return result
        
        
# ========= PARSE ONE PRICE SECTION =========
def parse_sheet_price_section(
    google_sheet_url: str,
    sheet_name: str,
    section_title: str,
    *,
    gabarit_width_flag: Optional[bool] = None,
    width_mm: int = 0,
    fabric_name: str = "",
    gabarit_height_mm: int = 0,
) -> Dict:
    """
    UA: Парсить ТІЛЬКИ вибрану секцію на вказаному листі.
    EN: Parse ONLY selected section on given sheet.

    Структура як у "Фальш-ролети":
      - зверху блоку: заголовок "Фальш-ролети, біла система" тощо;
      - далі кілька пояснювальних рядків;
      - рядок заголовків з "Тканина", "Висота рулону / тканини", "Габаритна висота ...";
      - наступний рядок: ширинні смуги;
      - далі: тканини до пустого рядка або наступного заголовку-секції.
      Може бути кілька таких під-таблиць (коли ширини не вміщаються в один блок).
    """
    result = {}

    wb = _download_workbook(google_sheet_url, force_refresh=False)
    sheet_name_clean = (sheet_name or "").strip()
    if sheet_name_clean not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name_clean]

    all_sections = find_sections_by_headers(
        ws,
        title_prefix="",
        min_merged_width=0,
        search_cols=6,
        case_insensitive=True,
        sheet_name=sheet_name,
    )

    result["sheet_name"] = sheet_name or ""
    result["sections"] = all_sections or None
    
    cg = getConfigBySheetName(sheet_name_clean)
    if cg.gbDiffWidthMm:
        result["GbDiffWidthMm"] = cg.gbDiffWidthMm
    else:
        result["GbDiffWidthMm"] = 0
        
    result["exist_control_side"] = cg.exist_control_side

    if not section_title:
        return result

    wanted_norm = _norm_title(section_title)
    target = next(
        (s for s in all_sections if _norm_title(s["title"]) == wanted_norm),
        None,
    )
    if not target:
        raise ValueError(f"Section '{section_title}' not found on sheet '{sheet_name}'")

    sections_sorted = sorted(all_sections, key=lambda x: (x["row"], x["col"]))

    idx_section = sections_sorted.index(target)
    start_row = target["row"]
    end_row = (
        sections_sorted[idx_section + 1]["row"] - 1
        if idx_section + 1 < len(sections_sorted)
        else ws.max_row
    )

    # 2) Find ALL header rows inside section
    #    (support one or multiple header blocks "Тканина / Висота / Габарит...")
    header_rows: List[int] = []
    for r in range(start_row, min(end_row, ws.max_row) + 1):
        vals = _row_values(ws, r)
        joined = " ".join([x or "" for x in vals]).lower()
        if "тканина" in joined and "висота" in joined and "габарит" in joined:
            header_rows.append(r)

    if not header_rows:
        raise RuntimeError(
            f"Header row not found in section '{section_title}' on sheet '{sheet_name}'"
        )

    # Для сумісності далі використовуємо перший заголовок як "основний"
    header_rows = sorted(header_rows)
    header_row = header_rows[0]

    # 3) Determine width bands and fabrics (support multiple sub-tables)
    #    Ми НЕ дублюємо парсинг: кожен рядок читається рівно один раз,
    #    а ціни по одній тканині з різних під-таблиць просто додаються в кінець.
    all_width_bands: List[Any] = []

    fabrics_map: Dict[str, Dict] = {}  # key: name.lower()
    fabric_order: List[str] = []  # to preserve first-seen order

    for i_hr, hr in enumerate(header_rows):
        header_vals = _row_values(ws, hr)

        # Find the first "ширина" column in this header block
        width_hdr_idx = None
        for i, v in enumerate(header_vals):
            if isinstance(v, str) and "ширина" in v.lower():
                width_hdr_idx = i
                break
        if width_hdr_idx is None:
            # Fallback: after third column
            width_hdr_idx = 3

        width_row = hr + 1
        width_row_vals = _row_values(ws, width_row)
        width_bands_part = [v for v in width_row_vals[width_hdr_idx:] if v]
        all_width_bands.extend(width_bands_part)

        # границя поточної під-таблиці: до наступного header_row або end_row
        next_header_row = (
            header_rows[i_hr + 1] if i_hr + 1 < len(header_rows) else end_row + 1
        )

        # 5) Fabrics list for this sub-table
        r = width_row + 1
        while r < next_header_row and r <= end_row:
            vals = _row_values(ws, r)

            # порожній рядок — кінець даних поточної під-таблиці
            if not any(vals):
                break

            name = (vals[0] or "").strip() if len(vals) > 0 else ""
            if not name:
                r += 1
                continue

            roll_h = _to_decimal(vals[1]) if len(vals) > 1 else None
            gabarit_limit = _to_decimal(vals[2]) if len(vals) > 2 else None

            price_cells = vals[width_hdr_idx:]
            prices_part = [
                (_to_decimal(pc) if _to_decimal(pc) is not None else None)
                for pc in price_cells
            ]

            key = name.lower()

            if key not in fabrics_map:
                fabrics_map[key] = {
                    "name": name,
                    "roll_height_mm": int(roll_h) if roll_h is not None else None,
                    "gabarit_limit_mm": (
                        int(gabarit_limit) if gabarit_limit is not None else None
                    ),
                    "prices_by_band": prices_part,
                }
                fabric_order.append(key)
            else:
                f = fabrics_map[key]
                # дозаповнюємо висоту/габарит, якщо раніше були None
                if f["roll_height_mm"] is None and roll_h is not None:
                    f["roll_height_mm"] = int(roll_h)
                if f["gabarit_limit_mm"] is None and gabarit_limit is not None:
                    f["gabarit_limit_mm"] = int(gabarit_limit)
                # додаємо нові ціни в кінець діапазону
                f["prices_by_band"].extend(prices_part)

            r += 1

    fabrics: List[Dict] = [fabrics_map[k] for k in fabric_order]
    width_bands = all_width_bands

    result["section_title"] = section_title or ""
    result["fabrics"] = fabrics or None
    result["section"] = target or ""
    result["GbDiffWidthMm"] = cg.gbDiffWidthMm or 0
    
    result = fillOptions(sheet_name, result, ws, header_row, section_title=section_title)

    if not width_mm or not gabarit_height_mm:
        return result

    real_width_mm = width_mm
    gb_width_mm = width_mm
    if cg.gbDiffWidthMm:
        # UA: У прайсі ширинні смуги йдуть по ширині тканини.
        # Якщо галочка "Габаритна ширина" УВІМКНЕНА, користувач вводить габаритну ширину,
        # тому для пошуку ціни віднімаємо різницю. А для опцій (за м.п.) зберігаємо саме габарит.
        # Якщо галочка ВИМКНЕНА, користувач вводить ширину по тканині, а габарит = тканина + різниця.
        diff = int(cg.gbDiffWidthMm)
        if gabarit_width_flag:
            real_width_mm = width_mm - diff
            gb_width_mm = width_mm - diff
        else:
            gb_width_mm = width_mm

    fabric = next(
        (f for f in fabrics if f["name"].lower() == fabric_name.lower()),
        None,
    )
    if not fabric:
        raise ValueError("Тканину не знайдено у вибраній секції")

    idx = pick_width_band(width_bands, real_width_mm)
    if idx is None:
        raise _width_out_of_range_error(fabric.get("gabarit_limit_mm"), fabric.get("roll_height_mm"))

    base_cell = fabric["prices_by_band"][idx]
    detail = _compute_price_detail(
        fabric=fabric,
        width_mm=gb_width_mm,
        gabarit_height_mm=gabarit_height_mm,
        bands=width_bands,
        band_idx=idx,
        base_cell=base_cell,
    )

    result["gabarit_limit_mm"] = detail["gabarit_limit_mm"]
    result["roll_height_mm"] = detail["roll_height_mm"]
    result["gb_width_mm"] = gb_width_mm or None
    result["band_index"] = detail["band_index"]
    result["band_label"] = detail["band_label"]
    result["base_price_eur"] = detail["base_price_eur"]
    result["surcharge_height_eur"] = detail["surcharge_height_eur"]

    return result


# ========= PRICE PREVIEW / FABRIC PARAMS =========


def price_preview_section(
    google_sheet_url: str,
    sheet_name: str,
    section_title: str,
    fabric_name: str,
    width_mm: int,
    gabarit_height_mm: int,
    gabarit_width_flag: bool,
) -> Dict:
    """
    EN: Calculate price preview for given fabric and dimensions.
    UA: Розрахунок прев'ю ціни для заданої тканини та розмірів.
    """
    parsed = parse_sheet_price_section(
        google_sheet_url=google_sheet_url,
        sheet_name=sheet_name,
        section_title=section_title,
        width_mm=width_mm,
        gabarit_height_mm=gabarit_height_mm,
        gabarit_width_flag=gabarit_width_flag,
    )

    width_for_price = width_mm  # hook for future width_input_dim logic

    fabrics = parsed.get("fabrics") or []
    fabric = next(
        (f for f in fabrics if f["name"].lower() == fabric_name.lower()),
        None,
    )
    if not fabric:
        raise ValueError("Тканину не знайдено у вибраній секції")

    bands = parsed["width_bands"]
    idx = pick_width_band(bands, width_for_price)
    if idx is None:
        raise _width_out_of_range_error(fabric.get("gabarit_limit_mm"), fabric.get("roll_height_mm"))

    base_cell = fabric["prices_by_band"][idx]
    magnets_price = parsed.get("magnets_price_eur") or Q("0.00")

    detail = _compute_price_detail(
        fabric=fabric,
        width_mm=width_mm,
        gabarit_height_mm=gabarit_height_mm,
        bands=bands,
        band_idx=idx,
        base_cell=base_cell,
        magnets_price=magnets_price,
    )

    return {
        "roll_height_mm": detail["roll_height_mm"],
        "gabarit_limit_mm": detail["gabarit_limit_mm"],
        "band_index": detail["band_index"],
        "band_label": detail["band_label"],
        "base_price_eur": detail["base_price_eur"],
        "surcharge_height_eur": detail["surcharge_height_eur"],
        "magnets_price_eur": detail["magnets_price_eur"],
    }
    

# ========= COMPONENTS SHEET (Комплектація) =========


def parse_components_sheet(
    google_sheet_url: str,
    sheet_name: str = "Комплектація",
    *,
    force_refresh: bool = False,
) -> Dict[str, Any]:
    """
    EN: Parse simple components list from 'Комплектація' sheet.
    UA: Парсить простий список комплектуючих з аркуша 'Комплектація'.

    Очікувана структура аркуша:
        Найменування | Од. вим. | Колір | Вартість, Євро
        <name>       | <unit>   | <color> | <price>

    Повертає:
        {
          "sheet_name": "Комплектація",
          "items": [
             {
               "name": "...",
               "unit": "шт",
               "color": "Білий",
               "price_eur": "2.199",
             },
             ...
          ],
          "names": [... унікальні найменування ...],
          "units": [... унікальні одиниці виміру ...],
          "colors": [... унікальні кольори ...],
        }
    """
    wb = _download_workbook(google_sheet_url, force_refresh=force_refresh)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]

    # 1) Знаходимо рядок заголовків
    header_row = None
    for r in range(1, ws.max_row + 1):
        vals = _row_values(ws, r)
        joined = " ".join([str(v) for v in vals if v]).strip().lower()
        if not joined:
            continue

        # шукаємо одночасно "найменування" і "вартість"
        if "найменування" in joined and "варт" in joined:
            header_row = r
            break

    if header_row is None:
        raise RuntimeError(
            f"Header row not found on sheet '{sheet_name}' (expect 'Найменування' / 'Вартість')"
        )

    header_vals = _row_values(ws, header_row)

    def _find_col(header_candidates) -> Optional[int]:
        """
        EN: Find column index by header text candidates.
        UA: Знаходить індекс колонки за можливими варіантами заголовка.
        """
        for idx, val in enumerate(header_vals):
            if not isinstance(val, str):
                continue
            norm = val.strip().lower()
            for cand in header_candidates:
                if cand in norm:
                    return idx
        return None

    # очікувані колонки
    name_idx = _find_col(["найменування"])
    unit_idx = _find_col(["од.", "од. вим", "од.вим"])
    color_idx = _find_col(["колір"])
    price_idx = _find_col(["вартість", "варт.", "ціна"])

    # як fallback — стандартний порядок колонок: 0,1,2,3
    if name_idx is None:
        name_idx = 0
    if unit_idx is None:
        unit_idx = 1
    if color_idx is None:
        color_idx = 2
    if price_idx is None:
        price_idx = 3

    items: List[Dict[str, Any]] = []
    names_set = set()
    units_set = set()
    colors_set = set()

    # 2) Читаємо всі рядки після заголовка
    for r in range(header_row + 1, ws.max_row + 1):
        vals = _row_values(ws, r)
        if not any(vals):
            # порожній рядок вважаємо кінцем таблиці
            continue

        name = (vals[name_idx] or "").strip() if len(vals) > name_idx else ""
        if not name:
            # пропускаємо рядки без найменування
            continue

        unit = (vals[unit_idx] or "").strip() if len(vals) > unit_idx else ""
        color = (vals[color_idx] or "").strip() if len(vals) > color_idx else ""

        raw_price = vals[price_idx] if len(vals) > price_idx else None
        price_dec = _to_decimal(raw_price)

        if price_dec is None:
            # якщо немає ціни — такий рядок не віддаємо на фронт
            logger.warning(
                "parse_components_sheet: empty/invalid price on row %s (name=%r)",
                r,
                name,
            )
            continue

        item = {
            "name": name,
            "unit": unit,
            "color": color,
            "price_eur": str(price_dec),
        }
        items.append(item)

        names_set.add(name)
        if unit:
            units_set.add(unit)
        if color:
            colors_set.add(color)

    result: Dict[str, Any] = {
        "sheet_name": sheet_name,
        "items": items,
        "names": sorted(names_set),
        "units": sorted(units_set),
        "colors": sorted(colors_set),
    }
    return result


def parse_fabrics_sheet(
    google_sheet_url: str,
    sheet_name: str = "Тканини до ролет",
    *,
    force_refresh: bool = False,
) -> Dict[str, Any]:
    wb = _download_workbook(google_sheet_url, force_refresh=force_refresh)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]

    header_row = None
    for r in range(1, ws.max_row + 1):
        vals = _row_values(ws, r)
        joined = " ".join([str(v) for v in vals if v]).strip().lower()
        if not joined:
            continue
        if "тканина" in joined and "ширина" in joined and "варт" in joined:
            header_row = r
            break

    if header_row is None:
        raise RuntimeError(
            f"Header row not found on sheet '{sheet_name}' (expect 'Тканина' / 'Ширина' / 'Вартість')"
        )

    header_vals = _row_values(ws, header_row)

    # Fixed columns: A-D (1-4) per spec to avoid ambiguous header matches.
    col_name, col_width, col_height, col_price = 0, 1, 2, 3
    if len(header_vals) < 4:
        raise RuntimeError("Недостатньо колонок у заголовку тканин (очікується 4).")

    extra_cut_label = ""
    extra_cut_price = None
    info_lines: List[str] = []
    for r in range(2, header_row):
        vals = _row_values(ws, r)
        text = " ".join([str(v) for v in vals[:4] if v]).strip()
        if text:
            info_lines.append(text)
        price_val = _to_decimal(vals[3]) if len(vals) > 3 else None
        if price_val is not None:
            extra_cut_price = price_val
            extra_cut_label = text

    items: List[Dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = _row_values(ws, r)
        if not any(vals):
            break
        name = (vals[col_name] or "").strip() if col_name < len(vals) else ""
        if not name:
            continue
        width_mm = _to_decimal(vals[col_width]) if col_width < len(vals) else None
        height_mm = _to_decimal(vals[col_height]) if col_height < len(vals) else None
        price_eur = _to_decimal(vals[col_price]) if col_price < len(vals) else None
        items.append(
            {
                "name": name,
                "roll_width_mm": int(width_mm) if width_mm is not None else None,
                "included_height_mm": int(height_mm) if height_mm is not None else None,
                "price_eur_mp": str(price_eur) if price_eur is not None else "0",
            }
        )

    return {
        "sheet_name": sheet_name,
        "items": items,
        "extra_cut_label": extra_cut_label or "",
        "extra_cut_price_eur": str(extra_cut_price or Q("0.00")),
        "info_lines": info_lines,
    }


def _mosquito_min_area_for_product(product_type: str) -> Q:
    name = (product_type or "").strip().lower()
    if "плісе двочаст" in name:
        return Q("4.0")
    if "дверні" in name or "посилені" in name or "плісе" in name:
        return Q("1.2")
    return Q("0.7")


def _mosquito_dimension_labels_for_product(product_type: str) -> Dict[str, str]:
    name = (product_type or "").strip().lower()
    if "анвіс" in name or "внутр. кріпл" in name:
        return {
            "width": "Ширина отвору",
            "height": "Висота отвору",
        }
    return {
        "width": "Ширина габаритна",
        "height": "Висота габаритна",
    }


def _mosquito_product_group(product_type: str) -> str:
    name = (product_type or "").strip().lower()
    if "внутрішні 10*30" in name:
        return "Внутрішні 10*30"
    if "зовнішні 10*20" in name:
        return "Зовнішні 10*20"
    if "дверні посилені" in name:
        return "Дверні посилені 14*40"
    if "дверні 17*25" in name:
        return "Дверні 17*25"
    if "ролетні внутр." in name:
        return "Ролетні внутрішнього кріплення"
    if "ролетні" in name:
        return "Ролетні"
    if "плісе двочаст" in name:
        return "Плісе двочастні"
    if "плісе одночаст" in name:
        return "Плісе одночастні"
    return product_type.strip()


def _mosquito_height_range_for_product(product_type: str):
    name = (product_type or "").strip().lower()
    if "до 1600" in name:
        return 0, 1600
    if "від 1600" in name and "до 2300" in name:
        return 1601, 2300
    return None, None


def select_mosquito_catalog_product(
    products: List[Dict[str, Any]],
    product_type: str,
    color: str,
    height_mm: int,
):
    candidates = [
        row for row in (products or [])
        if (
            (row.get("product_type") or "").strip().lower() == (product_type or "").strip().lower()
            or (row.get("raw_product_type") or "").strip().lower() == (product_type or "").strip().lower()
        )
        and (row.get("color") or "").strip().lower() == (color or "").strip().lower()
    ]
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    height_val = int(height_mm or 0)
    for row in candidates:
        min_h = row.get("height_range_min")
        max_h = row.get("height_range_max")
        if min_h is None and max_h is None:
            return row
        if min_h is not None and height_val < int(min_h):
            continue
        if max_h is not None and height_val > int(max_h):
            continue
        return row
    return candidates[-1]


def parse_mosquito_price_sheet(
    google_sheet_url: str,
    sheet_name: str = "Прайс",
    *,
    force_refresh: bool = False,
) -> Dict[str, Any]:
    wb = _download_workbook(google_sheet_url, force_refresh=force_refresh)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]
    info_lines: List[str] = []
    for r in range(1, 6):
        vals = _row_values(ws, r)
        text = " ".join([str(v) for v in vals if v]).strip()
        if text:
            info_lines.append(text)

    header_vals = _row_values(ws, 6)
    mesh_headers = [str(v).strip() for v in header_vals[2:6] if v]
    if len(mesh_headers) < 1:
        raise RuntimeError("Не знайдено заголовки полотен у прайсі москітних сіток.")

    products: List[Dict[str, Any]] = []
    current_product_type = ""
    for r in range(7, ws.max_row + 1):
        vals = _row_values(ws, r)
        if not any(vals):
            break
        product_type = (vals[0] or "").strip() if len(vals) > 0 and vals[0] else ""
        color = (vals[1] or "").strip() if len(vals) > 1 and vals[1] else ""
        if product_type:
            current_product_type = product_type
        if not current_product_type or not color:
            continue

        mesh_prices: Dict[str, str] = {}
        available_mesh_types: List[str] = []
        for idx, mesh_name in enumerate(mesh_headers, start=2):
            price = _to_decimal(vals[idx]) if len(vals) > idx else None
            if price is None:
                continue
            mesh_prices[mesh_name] = str(price)
            available_mesh_types.append(mesh_name)

        products.append(
            {
                "product_type": _mosquito_product_group(current_product_type),
                "raw_product_type": current_product_type,
                "color": color,
                "mesh_prices_usd_sqm": mesh_prices,
                "available_mesh_types": available_mesh_types,
                "min_area_sqm": str(_mosquito_min_area_for_product(current_product_type)),
                "dimension_labels": _mosquito_dimension_labels_for_product(current_product_type),
                "fiberglass_only": "ролетні" in current_product_type.lower(),
                "requires_sliding_side": "плісе" in current_product_type.lower(),
                "height_range_min": _mosquito_height_range_for_product(current_product_type)[0],
                "height_range_max": _mosquito_height_range_for_product(current_product_type)[1],
            }
        )

    options: List[Dict[str, Any]] = []
    for r in range(39, ws.max_row + 1):
        vals = _row_values(ws, r)
        if not any(vals):
            break
        name = (vals[0] or "").strip() if len(vals) > 0 else ""
        unit = (vals[4] or "").strip() if len(vals) > 4 else ""
        price = _to_decimal(vals[5]) if len(vals) > 5 else None
        if not name or price is None:
            continue
        options.append(
            {
                "name": name,
                "unit": unit,
                "price_usd": str(price),
            }
        )

    product_types = sorted({item["product_type"] for item in products})
    colors_by_product: Dict[str, List[str]] = {}
    mesh_by_product: Dict[str, List[str]] = {}
    for item in products:
        colors_by_product.setdefault(item["product_type"], [])
        if item["color"] not in colors_by_product[item["product_type"]]:
            colors_by_product[item["product_type"]].append(item["color"])
        mesh_by_product.setdefault(item["product_type"], [])
        for mesh in item["available_mesh_types"]:
            if mesh not in mesh_by_product[item["product_type"]]:
                mesh_by_product[item["product_type"]].append(mesh)

    return {
        "sheet_name": sheet_name,
        "info_lines": info_lines,
        "mesh_types": mesh_headers,
        "product_types": product_types,
        "products": products,
        "colors_by_product": colors_by_product,
        "mesh_types_by_product": mesh_by_product,
        "options": options,
    }


def parse_mosquito_components_sheet(
    google_sheet_url: str,
    sheet_name: str = "Комплектація",
    *,
    force_refresh: bool = False,
) -> Dict[str, Any]:
    def _component_group(name: str) -> str:
        lowered = (name or "").strip().lower()
        if "рол мс" in lowered or "ролет" in lowered:
            return "Ролетні"
        if "10*30" in lowered or "анвіс" in lowered or "внутр." in lowered:
            return "Внутрішні 10*30"
        if "10*20" in lowered or "зовніш" in lowered:
            return "Зовнішні 10*20"
        if "дверн" in lowered and ("17*25" in lowered or "посилен" not in lowered):
            return "Дверні 17*25"
        if "посилен" in lowered or "14*40" in lowered or "17*40" in lowered:
            return "Дверні посилені 14*40"
        if "плісе" in lowered:
            return "Плісе"
        return "Усі"

    parsed = parse_components_sheet(
        google_sheet_url=google_sheet_url,
        sheet_name=sheet_name,
        force_refresh=force_refresh,
    )
    items = []
    groups = set()
    names_by_group: Dict[str, List[str]] = {}
    for item in parsed.get("items") or []:
        group = _component_group(item.get("name") or "")
        groups.add(group)
        names_by_group.setdefault(group, [])
        if (item.get("name") or "") not in names_by_group[group]:
            names_by_group[group].append(item.get("name") or "")
        items.append(
            {
                "name": item.get("name") or "",
                "color": item.get("color") or "",
                "unit": item.get("unit") or "",
                "price_usd": item.get("price_eur") or "0",
                "product_group": group,
            }
        )
    return {
        "sheet_name": parsed.get("sheet_name") or sheet_name,
        "items": items,
        "product_groups": sorted(groups),
        "names_by_group": names_by_group,
        "names": parsed.get("names") or [],
        "units": parsed.get("units") or [],
        "colors": parsed.get("colors") or [],
    }


def build_mosquito_warnings(
    product_type: str,
    mesh_type: str,
    width_mm: int,
    height_mm: int,
    area_sqm,
) -> List[str]:
    name = (product_type or "").strip().lower()
    mesh = (mesh_type or "").strip().lower()
    max_side = max(int(width_mm or 0), int(height_mm or 0))
    warnings: List[str] = []

    if ("10*30" in name or "10*20" in name):
        if "антикіт" in mesh:
            if max_side > 1400:
                warnings.append("Негарантійний виріб, рекомендується встановлення імпоста")
        elif max_side > 1700:
            warnings.append("Негарантійний виріб, рекомендується встановлення імпоста")

    if "17*25" in name:
        if "антикіт" in mesh:
            if max_side > 1600:
                warnings.append("Негарантійний виріб, рекомендується встановлення імпоста")
        elif max_side > 2000:
            warnings.append("Негарантійний виріб, рекомендується встановлення імпоста")

    if "ролетні" in name:
        if int(width_mm or 0) > 1500:
            warnings.append("Негарантійний виріб, рекомендується ширина до 1500мм")
        try:
            area_val = Q(str(area_sqm or 0))
        except Exception:
            area_val = Q("0")
        if area_val > Q("2.5"):
            warnings.append("Негарантійний виріб, рекомендується квадратура до 2,5 м.кв.")

    if "плісе" in name and "двочаст" not in name and int(width_mm or 0) >= 3000:
        warnings.append("Оберіть двочасну сітку")

    return warnings
